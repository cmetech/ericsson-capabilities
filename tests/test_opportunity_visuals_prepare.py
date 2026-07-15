import json
import os
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook


REPO = Path(__file__).resolve().parents[1]
SCRIPTS = REPO / "skills/ericsson/opportunity-visuals/scripts"
sys.path.insert(0, str(SCRIPTS))

from opportunity_data import (  # noqa: E402
    DataContractError,
    apply_filters,
    classify_transition,
    inspect_source,
    load_source,
    normalize_rank,
    normalize_rows,
    parse_month_header,
    resolve_mapping,
    select_records,
    validate_semantics,
)
import prepare_opportunities  # noqa: E402
from prepare_opportunities import analyze, prepare  # noqa: E402


HEADERS = [
    "Area",
    "Sub-area",
    "Opportunity Name",
    "TCV",
    "Probability",
    "Mar '26",
    "Apr '26",
    "May '26",
]
ROW = ["North", "Cloud", "Project Cedar", "Large", "High", "Ideation", "Proposal", "Won"]


@pytest.fixture
def csv_source(tmp_path):
    source = tmp_path / "pipeline.csv"
    source.write_text(
        ",".join(HEADERS) + "\n" + ",".join(ROW) + "\n",
        encoding="utf-8",
    )
    return source


@pytest.fixture
def json_source(tmp_path):
    source = tmp_path / "pipeline.json"
    source.write_text(json.dumps([dict(zip(HEADERS, ROW, strict=True))]), encoding="utf-8")
    return source


@pytest.fixture
def xlsx_source(tmp_path):
    source = tmp_path / "pipeline.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pipeline"
    sheet.append(HEADERS)
    sheet.append(ROW)
    workbook.save(source)
    return source


@pytest.fixture
def xlsx_with_formula(tmp_path):
    source = tmp_path / "formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pipeline"
    sheet.append(HEADERS)
    row = ROW.copy()
    row[3] = "=1+1"
    sheet.append(row)
    workbook.save(source)
    return source


def test_loaders_return_equivalent_rows(csv_source, json_source, xlsx_source):
    csv_rows, _ = load_source(csv_source, None)
    json_rows, _ = load_source(json_source, None)
    xlsx_rows, meta = load_source(xlsx_source, "Pipeline")
    assert csv_rows == json_rows == xlsx_rows
    assert meta["sheet"] == "Pipeline"


@pytest.mark.parametrize("constant", ["NaN", "Infinity", "-Infinity"])
def test_json_source_rejects_nonstandard_numeric_constants(tmp_path, constant):
    source = tmp_path / "nonstandard.json"
    source.write_text(
        '[{"Area":"A","Sub-area":"S","Opportunity Name":"O",'
        f'"TCV":1,"Probability":{constant},"Mar \'26":"A","Apr \'26":"B"}}]',
        encoding="utf-8",
    )

    with pytest.raises(DataContractError) as caught:
        load_source(source)

    assert caught.value.code == "invalid_json"


def test_loaders_and_mapping_preserve_padded_header_labels(tmp_path):
    headers = [
        "  Area  ",
        " Sub-area",
        "Opportunity Name ",
        " TCV ",
        "Probability  ",
        " Mar '26 ",
        " Apr '26 ",
        "Apr '26 Probability ",
    ]
    expected_rows = [dict(zip(headers, ROW, strict=True))]

    csv_source = tmp_path / "padded.csv"
    csv_source.write_text(
        ",".join(headers) + "\n" + ",".join(ROW) + "\n",
        encoding="utf-8",
    )
    json_source = tmp_path / "padded.json"
    json_source.write_text(json.dumps(expected_rows), encoding="utf-8")
    xlsx_source = tmp_path / "padded.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pipeline"
    sheet.append(headers)
    sheet.append(ROW)
    workbook.save(xlsx_source)

    csv_rows, _ = load_source(csv_source)
    json_rows, _ = load_source(json_source)
    xlsx_rows, _ = load_source(xlsx_source, "Pipeline")

    assert csv_rows == json_rows == xlsx_rows == expected_rows
    mapping = resolve_mapping(list(csv_rows[0]), None)
    assert mapping["area"] == "  Area  "
    assert mapping["sub_area"] == " Sub-area"
    assert mapping["months"] == [
        {"key": "2026-03", "label": " Mar '26 ", "stage": " Mar '26 "},
        {
            "key": "2026-04",
            "label": " Apr '26 ",
            "stage": " Apr '26 ",
            "probability": "Apr '26 Probability ",
        },
    ]


def test_inspect_lists_sheets_headers_and_month_candidates(xlsx_source):
    report = inspect_source(xlsx_source, "Pipeline")
    assert report["format"] == "xlsx"
    assert report["headers"][:5] == [
        "Area",
        "Sub-area",
        "Opportunity Name",
        "TCV",
        "Probability",
    ]
    assert [month["label"] for month in report["month_candidates"]] == [
        "Mar '26",
        "Apr '26",
        "May '26",
    ]
    assert report["row_count"] == 1
    assert report["source"] == "pipeline.xlsx"
    assert len(report["sha256"]) == 64
    assert "rows" not in report


def test_parse_month_header_handles_loop24_labels():
    assert parse_month_header("March’26") == ("2026-03", "stage")
    assert parse_month_header("Apr '26 Probability") == ("2026-04", "probability")
    assert parse_month_header("2026-05 Status") == ("2026-05", "stage")
    assert parse_month_header("Opportunity Name") is None


def test_parse_month_header_handles_four_digit_year_and_normalized_spacing():
    assert parse_month_header("September_2027 stage") == ("2027-09", "stage")
    assert parse_month_header("  2028-01   Probability ") == ("2028-01", "probability")


def test_ambiguous_alias_requires_explicit_mapping():
    with pytest.raises(DataContractError, match="ambiguous field area"):
        resolve_mapping(["Area", "Sales Area", "Name", "TCV", "Probability"], None)


def test_explicit_mapping_is_validated_and_months_are_sorted():
    headers = [
        "Sales Area",
        "Subarea",
        "Deal Name",
        "Total Contract Value",
        "Current Probability",
        "Apr '26",
        "Mar '26 Probability",
        "Mar '26",
    ]
    explicit = {
        "area": "Sales Area",
        "sub_area": "Subarea",
        "opportunity_name": "Deal Name",
        "tcv": "Total Contract Value",
        "probability": "Current Probability",
        "months": [
            {"key": "2026-04", "label": "April", "stage": "Apr '26"},
            {
                "key": "2026-03",
                "label": "March",
                "stage": "Mar '26",
                "probability": "Mar '26 Probability",
            },
        ],
    }
    mapping = resolve_mapping(headers, explicit)
    assert mapping["area"] == "Sales Area"
    assert mapping["months"] == [
        {
            "key": "2026-03",
            "label": "March",
            "stage": "Mar '26",
            "probability": "Mar '26 Probability",
        },
        {"key": "2026-04", "label": "April", "stage": "Apr '26"},
    ]


@pytest.mark.parametrize(
    ("months", "reused_header"),
    [
        (
            [
                {"key": "2026-03", "label": "March", "stage": "Mar '26"},
                {"key": "2026-04", "label": "April", "stage": "Mar '26"},
            ],
            "Mar '26",
        ),
        (
            [
                {
                    "key": "2026-03",
                    "label": "March",
                    "stage": "Mar '26",
                    "probability": "Mar '26",
                },
            ],
            "Mar '26",
        ),
        (
            [
                {
                    "key": "2026-03",
                    "label": "March",
                    "stage": "Mar '26",
                    "probability": "Monthly Probability",
                },
                {
                    "key": "2026-04",
                    "label": "April",
                    "stage": "Apr '26",
                    "probability": "Monthly Probability",
                },
            ],
            "Monthly Probability",
        ),
    ],
    ids=("stage-stage", "stage-probability", "probability-probability"),
)
def test_explicit_mapping_rejects_reused_month_headers(months, reused_header):
    headers = [
        "Area",
        "Sub-area",
        "Opportunity Name",
        "TCV",
        "Probability",
        "Mar '26",
        "Apr '26",
        "Monthly Probability",
    ]
    explicit = {
        "area": "Area",
        "sub_area": "Sub-area",
        "opportunity_name": "Opportunity Name",
        "tcv": "TCV",
        "probability": "Probability",
        "months": months,
    }

    with pytest.raises(DataContractError, match="reused month header") as error:
        resolve_mapping(headers, explicit)
    assert error.value.code == "invalid_mapping"
    assert error.value.details["header"] == reused_header


def test_auto_mapping_pairs_month_probability_and_preserves_labels():
    headers = [
        "Area",
        "Sub-area",
        "Opportunity Name",
        "TCV",
        "Probability",
        "Apr '26 Probability",
        "Mar '26",
        "Apr '26",
    ]
    mapping = resolve_mapping(headers, None)
    assert mapping["months"] == [
        {"key": "2026-03", "label": "Mar '26", "stage": "Mar '26"},
        {
            "key": "2026-04",
            "label": "Apr '26",
            "stage": "Apr '26",
            "probability": "Apr '26 Probability",
        },
    ]


def test_duplicate_stage_month_requires_explicit_mapping():
    headers = [
        "Area",
        "Sub-area",
        "Opportunity Name",
        "TCV",
        "Probability",
        "Mar '26",
        "March 2026 Status",
    ]
    with pytest.raises(DataContractError, match="ambiguous stage month 2026-03"):
        resolve_mapping(headers, None)


def test_json_with_multiple_arrays_requires_key(tmp_path):
    source = tmp_path / "pipeline.json"
    source.write_text(json.dumps({"current": [{"Area": "A"}], "archive": [{"Area": "B"}]}))
    with pytest.raises(DataContractError, match="JSON array key is required") as error:
        load_source(source, None, None)
    assert error.value.details == {"array_keys": ["current", "archive"]}
    rows, meta = load_source(source, None, "current")
    assert rows == [{"Area": "A"}]
    assert meta["json_key"] == "current"


def test_xlsx_formula_without_cached_value_is_reported(xlsx_with_formula):
    _, meta = load_source(xlsx_with_formula, "Pipeline", None)
    assert meta["uncached_formulas"] == [{"row": 2, "header": "TCV"}]


@pytest.mark.parametrize(
    ("name", "contents", "message"),
    [
        ("empty.csv", "", "source contains no rows"),
        ("blank.csv", "Area,,TCV\nA,B,Large\n", "blank header"),
        ("duplicate.csv", "Area,Area,TCV\nA,B,Large\n", "duplicate header"),
        ("short.csv", "Area,TCV\nA\n", "row 2 does not match headers"),
    ],
)
def test_csv_rejects_invalid_table_shapes(tmp_path, name, contents, message):
    source = tmp_path / name
    source.write_text(contents, encoding="utf-8")
    with pytest.raises(DataContractError, match=message):
        load_source(source)


def test_json_rejects_rows_with_different_headers(tmp_path):
    source = tmp_path / "pipeline.json"
    source.write_text(json.dumps([{"Area": "A", "TCV": "Large"}, {"Area": "B"}]))
    with pytest.raises(DataContractError, match="row 2 does not match headers"):
        load_source(source)


def test_inspect_cli_prints_one_safe_json_object_and_writes_nothing(csv_source):
    before = sorted(path.name for path in csv_source.parent.iterdir())
    result = subprocess.run(
        [sys.executable, str(SCRIPTS / "prepare_opportunities.py"), "inspect", str(csv_source)],
        check=False,
        capture_output=True,
        text=True,
    )
    after = sorted(path.name for path in csv_source.parent.iterdir())

    assert result.returncode == 0
    assert result.stderr == ""
    report = json.loads(result.stdout)
    assert report["ok"] is True
    assert report["format"] == "csv"
    assert result.stdout.count("\n") == 1
    assert "Project Cedar" not in result.stdout
    assert before == after


def test_inspect_cli_returns_structured_error_for_multi_array_json(tmp_path):
    source = tmp_path / "pipeline.json"
    source.write_text(json.dumps({"current": [{"Area": "A"}], "archive": [{"Area": "B"}]}))
    result = subprocess.run(
        [sys.executable, str(SCRIPTS / "prepare_opportunities.py"), "inspect", str(source)],
        check=False,
        capture_output=True,
        text=True,
    )
    assert result.returncode == 2
    assert json.loads(result.stdout) == {
        "ok": False,
        "error": {
            "code": "json_key_required",
            "message": "JSON array key is required",
            "details": {"array_keys": ["current", "archive"]},
        },
    }


@pytest.fixture
def semantics():
    return {
        "positive_terminals": ["Won"],
        "negative_terminals": ["Lost", "Cancelled"],
        "non_terminal_stages": ["Discovery"],
        "stage_paths": [
            ["Ideation", "Solution", "Proposal", "SDP2", "Won"],
            ["POC", "Workshop", "Commercials", "Won"],
        ],
        "positive_transitions": [["Proposal", "Workshop"]],
        "tcv_order": ["X-Small", "Small", "Medium", "Large", "X-Large"],
        "probability_order": ["Low", "Medium", "High", "Certain"],
    }


@pytest.mark.parametrize(
    ("previous", "current", "expected"),
    [
        ({"stage": "Ideation"}, {"stage": "Proposal"}, "positive"),
        ({"stage": "Proposal"}, {"stage": "Workshop"}, "positive"),
        ({"stage": "Solution"}, {"stage": "Ideation"}, "negative"),
        ({"stage": "Solution"}, {"stage": "Solution"}, "neutral"),
        ({"stage": "Proposal"}, {"stage": "Won"}, "won"),
        ({"stage": "Proposal"}, {"stage": "Lost"}, "lost"),
        (
            {"stage": "Solution", "probability_sort": 3},
            {"stage": "Proposal", "probability_sort": 2},
            "mixed",
        ),
        ({"stage": "Discovery"}, {"stage": "Deferred"}, "unknown"),
    ],
)
def test_classify_transition(previous, current, expected, semantics):
    assert classify_transition(previous, current, semantics) == expected


@pytest.mark.parametrize(
    ("value", "field", "expected"),
    [
        ("$2.5M", "tcv", ("$2.5M", 2_500_000.0, "numeric")),
        ("1,250", "tcv", ("1,250", 1250.0, "numeric")),
        (0.25, "probability", ("0.25", 25.0, "numeric")),
        ("75%", "probability", ("75%", 75.0, "numeric")),
        ("high", "probability", ("high", 2, "categorical")),
    ],
)
def test_normalize_rank_accepts_numeric_and_categorical_values(
    value, field, expected, semantics
):
    order = semantics[f"{field}_order"]
    assert normalize_rank(value, order, field) == expected


@pytest.mark.parametrize("value", [-0.01, -1, 101, "101%", "-0.5%"])
def test_normalize_rank_rejects_numeric_probability_outside_percentage_range(value):
    with pytest.raises(DataContractError) as caught:
        normalize_rank(value, [], "probability")

    assert caught.value.code == "invalid_probability"


@pytest.mark.parametrize("field", ["tcv", "probability"])
@pytest.mark.parametrize("value", [10**1000, "9" * 1001])
def test_normalize_rank_rejects_oversized_numeric_conversions_stably(
    value, field
):
    with pytest.raises(DataContractError) as caught:
        normalize_rank(value, [], field)

    assert caught.value.code == f"invalid_{field}"


def test_oversized_monthly_probability_is_a_row_local_exclusion(semantics):
    row = {
        "ID": "HUGE-MONTHLY",
        "Area": "Core",
        "Sub-area": "Cloud",
        "Opportunity Name": "Oversized probability",
        "TCV": "$1M",
        "Probability": 50,
        "Mar '26": "Solution",
        "Mar '26 Probability": 50,
        "Apr '26": "Proposal",
        "Apr '26 Probability": 10**1000,
        "May '26": "Proposal",
        "May '26 Probability": 50,
        "Jun '26": "Proposal",
        "Jun '26 Probability": 50,
    }

    records, exclusions = normalize_rows([row], _showcase_mapping(), semantics)

    assert records == []
    assert exclusions[0]["code"] == "invalid_probability"


def test_invalid_monthly_probability_is_excluded_before_it_can_change_classification(
    semantics,
):
    row = {
        "ID": "BAD-MONTHLY",
        "Area": "Core",
        "Sub-area": "Cloud",
        "Opportunity Name": "Invalid probability",
        "TCV": "$1M",
        "Probability": 50,
        "Mar '26": "Solution",
        "Mar '26 Probability": 50,
        "Apr '26": "Proposal",
        "Apr '26 Probability": 101,
        "May '26": "Proposal",
        "May '26 Probability": 50,
        "Jun '26": "Proposal",
        "Jun '26 Probability": 50,
    }

    records, exclusions = normalize_rows([row], _showcase_mapping(), semantics)

    assert records == []
    assert exclusions == [
        {
            "id": "BAD-MONTHLY",
            "source_row": 2,
            "code": "invalid_probability",
            "message": "Invalid probability value: 101",
        }
    ]


def _showcase_rows():
    histories = {
        "OV-001": [("Ideation", "Low"), ("Proposal", "Medium"), ("Won", "Certain"), ("In Delivery", "Certain")],
        "OV-002": [("Proposal", "Medium"), ("Workshop", "High"), ("Commercials", "High"), ("Commercials", "High")],
        "OV-003": [("Solution", "High"), ("Ideation", "Medium"), ("Lost", "Low"), ("", "")],
        "OV-004": [("Solution", "Medium"), ("Solution", "Medium"), ("Solution", "Medium"), ("Solution", "Medium")],
        "OV-005": [("", ""), ("", ""), ("Discovery", "Low"), ("", "")],
        "OV-006": [("Ideation", "Low"), ("", ""), ("Proposal", "High"), ("Proposal", "High")],
        "OV-007": [("Solution", "High"), ("Proposal", "Medium"), ("Proposal", "Medium"), ("Proposal", "Medium")],
        "OV-008": [("Ideation", "Low"), ("Solution", "Medium"), ("Solution", "Medium"), ("Solution", "Medium")],
        "OV-009": [("Proposal", "High"), ("Lost", "Low"), ("Restarted", "Medium"), ("Won", "Certain")],
        "OV-010": [("Proposal", "High"), ("Solution", "Medium"), ("Proposal", "High"), ("Won", "Certain")],
        "OV-011": [("Discovery", "Low"), ("Deferred", "Low"), ("Deferred", "Low"), ("Deferred", "Low")],
        "OV-012": [("Ideation", "Low"), ("Solution", "Medium"), ("Proposal", "High"), ("Proposal", "High")],
    }
    fixed = {
        "OV-001": ("Core Group", "Core", "Aurora Core Renewal", "X-Large", "Certain"),
        "OV-002": ("Core Group", "Automation", "Beacon Automation", "Large", "High"),
        "OV-003": ("Cloud Group", "Assurance", "Cedar Assurance", "X-Large", "Low"),
        "OV-004": ("Cloud Group", "Assurance", "Delta Capacity", "Medium", "Medium"),
        "OV-005": ("Edge Group", "Discovery", "Echo Modernization", "Small", "Low"),
        "OV-006": ("Core Group", "Automation", "Fjord Analytics", "Medium", "High"),
        "OV-007": ("OSS Group", "Orchestration", "Grove Orchestration", "Large", "Medium"),
        "OV-008": ("OSS Group", "Observability", "Harbor Observability <Pilot> =1+1", "Small", "Medium"),
        "OV-009": ("Edge Group", "Delivery", "Ion Edge Program", "X-Large", "Low"),
        "OV-010": ("Core Group", "Core", "Juniper Expansion", "Large", "Certain"),
        "OV-011": ("Edge Group", "Discovery", "Kite Discovery", "X-Small", "Low"),
        "OV-012": ("OSS Group", "Platform", "Lumen Platform", "Medium", "High"),
    }
    rows = []
    for row_id, fields in fixed.items():
        area, sub_area, name, tcv, probability = fields
        row = {
            "ID": row_id,
            "Area": area,
            "Sub-area": sub_area,
            "Opportunity Name": name,
            "TCV": tcv,
            "Probability": probability,
        }
        for label, values in zip(("Mar '26", "Apr '26", "May '26", "Jun '26"), histories[row_id], strict=True):
            row[label], row[f"{label} Probability"] = values
        rows.append(row)
    return rows


def _showcase_mapping():
    return {
        "area": "Area",
        "sub_area": "Sub-area",
        "opportunity_name": "Opportunity Name",
        "tcv": "TCV",
        "probability": "Probability",
        "months": [
            {
                "key": f"2026-{month:02d}",
                "label": label,
                "stage": label,
                "probability": f"{label} Probability",
            }
            for month, label in enumerate(
                ("Mar '26", "Apr '26", "May '26", "Jun '26"), start=3
            )
        ],
    }


def test_normalize_rows_preserves_values_blanks_ranks_and_warnings(semantics):
    records, excluded = normalize_rows(_showcase_rows(), _showcase_mapping(), semantics)
    by_id = {record["id"]: record for record in records}

    assert "OV-005" not in by_id
    assert excluded == [
        {
            "id": "OV-005",
            "source_row": 6,
            "code": "insufficient_stages",
            "message": "At least two populated stages are required",
        }
    ]
    assert by_id["OV-008"]["opportunity_name"] == "Harbor Observability <Pilot> =1+1"
    assert by_id["OV-012"]["tcv"] == {
        "display": "Medium",
        "sort": 2,
        "kind": "categorical",
    }
    assert by_id["OV-012"]["probability"] == {
        "display": "High",
        "sort": 2,
        "kind": "categorical",
    }
    assert [month["classification"] for month in by_id["OV-006"]["months"]] == [
        "initial",
        "empty",
        "positive",
        "neutral",
    ]
    assert by_id["OV-006"]["months"][2]["skipped_months"] == ["Apr '26"]
    assert "skipped_blank_months" in {
        warning["code"] for warning in by_id["OV-006"]["warnings"]
    }
    assert [month["classification"] for month in by_id["OV-007"]["months"]] == [
        "initial",
        "mixed",
        "neutral",
        "neutral",
    ]
    assert "mixed_signals" in {warning["code"] for warning in by_id["OV-007"]["warnings"]}
    assert "unknown_transition" in {
        warning["code"] for warning in by_id["OV-011"]["warnings"]
    }
    assert len(by_id["OV-001"]["months"]) == 3
    assert by_id["OV-001"]["terminal"] == {
        "kind": "positive",
        "month": "2026-05",
        "index": 2,
    }


def test_normalize_rows_generates_unique_fallback_ids_for_json_null_ids(semantics):
    rows = [dict(_showcase_rows()[0], ID=None), dict(_showcase_rows()[1], ID=None)]

    records, exclusions = normalize_rows(rows, _showcase_mapping(), semantics)

    assert exclusions == []
    assert [record["id"] for record in records] == ["ROW-0002", "ROW-0003"]


def test_normalize_rows_generates_unique_fallback_ids_for_xlsx_blank_ids(
    tmp_path, semantics
):
    rows = [dict(_showcase_rows()[0], ID=None), dict(_showcase_rows()[1], ID=None)]
    source = tmp_path / "blank-ids.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pipeline"
    headers = list(rows[0])
    sheet.append(headers)
    for row in rows:
        sheet.append([row[header] for header in headers])
    workbook.save(source)

    loaded, _ = load_source(source, "Pipeline")
    records, exclusions = normalize_rows(loaded, _showcase_mapping(), semantics)

    assert exclusions == []
    assert [record["id"] for record in records] == ["ROW-0002", "ROW-0003"]


def test_pre_range_exclusions_use_fallback_ids_for_blank_explicit_ids(semantics):
    rows = [dict(_showcase_rows()[0], ID=None), dict(_showcase_rows()[1], ID="")]
    exclusions = prepare_opportunities._terminal_before_range_exclusions(
        rows,
        _showcase_mapping()["months"],
        3,
        semantics,
    )

    assert [item["id"] for item in exclusions] == ["ROW-0002"]


def test_normalize_rows_rejects_duplicate_explicit_nonblank_ids(semantics):
    rows = [dict(_showcase_rows()[0], ID="DUPLICATE"), dict(_showcase_rows()[1], ID="DUPLICATE")]

    with pytest.raises(DataContractError) as caught:
        normalize_rows(rows, _showcase_mapping(), semantics)

    assert caught.value.code == "duplicate_id"


def test_select_records_applies_each_view_and_deterministic_sorting(semantics):
    records, _ = normalize_rows(_showcase_rows(), _showcase_mapping(), semantics)

    assert [record["id"] for record in select_records(records, "wins")[0]] == [
        "OV-001",
        "OV-010",
    ]
    assert [record["id"] for record in select_records(records, "losses")[0]] == [
        "OV-003",
        "OV-009",
    ]
    all_ids = [record["id"] for record in select_records(records, "all-progression")[0]]
    positive_ids = [
        record["id"]
        for record in select_records(records, "positive-progression")[0]
    ]
    assert "OV-004" in all_ids
    assert "OV-004" not in positive_ids
    assert "OV-005" not in all_ids
    assert all_ids == [
        "OV-003",
        "OV-004",
        "OV-002",
        "OV-006",
        "OV-001",
        "OV-010",
        "OV-009",
        "OV-011",
        "OV-008",
        "OV-007",
        "OV-012",
    ]


def test_apply_filters_matches_case_insensitively_and_uses_rank_bounds(semantics):
    records, _ = normalize_rows(_showcase_rows(), _showcase_mapping(), semantics)
    filters = {
        "areas": ["core group"],
        "sub_areas": ["automation"],
        "opportunity_contains": "automation",
        "tcv_min": 2,
        "tcv_max": 4,
        "probability_min": 1,
        "probability_max": 3,
    }

    included, excluded = apply_filters(records, filters)

    assert [record["id"] for record in included] == ["OV-002"]
    assert all(record["area"] == "Core Group" for record in included)
    assert all(item["code"] == "filter_not_matched" for item in excluded)
    assert all("Harbor Observability" not in item["message"] for item in excluded)


@pytest.mark.parametrize("bound", [True, False, float("nan"), float("inf"), float("-inf")])
def test_apply_filters_rejects_boolean_and_nonfinite_numeric_bounds(bound):
    with pytest.raises(DataContractError) as caught:
        apply_filters([], {"probability_min": bound})

    assert caught.value.code == "invalid_filters"


@pytest.mark.parametrize("prefix", ["tcv", "probability"])
def test_apply_filters_rejects_oversized_integer_bounds_stably(prefix):
    with pytest.raises(DataContractError) as caught:
        apply_filters([], {f"{prefix}_min": 10**1000})

    assert caught.value.code == "invalid_filters"


def test_prepare_cli_writes_stable_artifacts_and_refuses_nonempty_output(
    tmp_path, semantics
):
    source = tmp_path / "pipeline.csv"
    rows = _showcase_rows()[:2]
    headers = list(rows[0])
    source.write_text(
        ",".join(headers)
        + "\n"
        + "\n".join(
            ",".join(str(row[header]) for header in headers) for row in rows
        )
        + "\n",
        encoding="utf-8",
    )
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")
    output_dir = tmp_path / "prepared"
    command = [
        sys.executable,
        str(SCRIPTS / "prepare_opportunities.py"),
        "prepare",
        str(source),
        "--view",
        "wins",
        "--semantics",
        str(semantics_path),
        "--output-dir",
        str(output_dir),
    ]

    result = subprocess.run(command, check=False, capture_output=True, text=True)

    assert result.returncode == 0, result.stdout + result.stderr
    assert result.stderr == ""
    assert sorted(path.name for path in output_dir.iterdir()) == [
        "exclusions.json",
        "normalized-data.json",
        "source-summary.json",
    ]
    summary = json.loads((output_dir / "source-summary.json").read_text())
    normalized = json.loads((output_dir / "normalized-data.json").read_text())
    assert summary["source"] == "pipeline.csv"
    assert len(summary["sha256"]) == 64
    assert normalized["source"] == {
        "basename": "pipeline.csv",
        "sha256": summary["sha256"],
        "sheet": None,
    }
    assert normalized["schema_version"] == 1
    assert [record["id"] for record in normalized["records"]] == ["OV-001"]
    assert json.loads((output_dir / "exclusions.json").read_text()) == {
        "exclusions": normalized["exclusions"]
    }
    for artifact in output_dir.iterdir():
        text = artifact.read_text(encoding="utf-8")
        assert text.endswith("\n")
        assert text == json.dumps(json.loads(text), indent=2, sort_keys=True) + "\n"

    repeated = subprocess.run(command, check=False, capture_output=True, text=True)
    assert repeated.returncode == 2
    assert json.loads(repeated.stdout)["error"]["code"] == "output_exists"


def test_prepare_function_rejects_formula_only_when_selected_field_is_uncached(
    tmp_path, xlsx_with_formula, semantics
):
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")

    with pytest.raises(DataContractError) as error:
        prepare(
            xlsx_with_formula,
            "wins",
            semantics_path,
            tmp_path / "prepared",
            sheet="Pipeline",
        )
    assert error.value.code == "formula_cache_missing"


def test_prepare_preserves_source_rows_after_terminal_before_selected_range(
    tmp_path, semantics
):
    source = tmp_path / "pipeline.csv"
    source.write_text(
        "ID,Area,Sub-area,Opportunity Name,TCV,Probability,Mar '26,Apr '26,May '26,Jun '26\n"
        "OV-A,Area A,Sub A,Terminal First,Large,High,Proposal,Lost,Restarted,Won\n"
        "OV-B,Area B,Sub B,Still Eligible,Medium,Medium,,,Ideation,Proposal\n",
        encoding="utf-8",
    )
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")
    output_dir = tmp_path / "prepared"

    prepare(
        source,
        "all-progression",
        semantics_path,
        output_dir,
        months=["May '26", "Jun '26"],
    )

    normalized = json.loads((output_dir / "normalized-data.json").read_text())
    assert [(record["id"], record["source_row"]) for record in normalized["records"]] == [
        ("OV-B", 3)
    ]
    assert normalized["exclusions"] == [
        {
            "id": "OV-A",
            "source_row": 2,
            "code": "terminal_before_range",
            "message": "Row reached a terminal before the selected range",
        }
    ]


def test_semantics_reject_conflicting_case_insensitive_configuration(semantics):
    conflicting = dict(semantics)
    conflicting["stage_paths"] = [["Ideation", "Solution"], ["solution", "Ideation"]]

    with pytest.raises(DataContractError) as error:
        classify_transition({"stage": "Ideation"}, {"stage": "Solution"}, conflicting)
    assert error.value.code == "conflicting_stage_order"


def test_semantics_reject_transitive_case_insensitive_cycle(semantics):
    conflicting = dict(semantics)
    conflicting["stage_paths"] = [["A", "B"], ["b", "C"], ["c", "a"]]

    with pytest.raises(DataContractError) as error:
        classify_transition({"stage": "A"}, {"stage": "B"}, conflicting)
    assert error.value.code == "conflicting_stage_order"
    assert str(error.value) == "Stage paths assign contradictory ranks"


def test_prepare_retains_warning_for_record_excluded_by_view(tmp_path, semantics):
    source = tmp_path / "pipeline.csv"
    source.write_text(
        "ID,Area,Sub-area,Opportunity Name,TCV,Probability,Mar '26,Apr '26\n"
        "OV-U,Area,Sub,Unknown Path,Large,High,Discovery,Deferred\n",
        encoding="utf-8",
    )
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")
    output_dir = tmp_path / "prepared"

    prepare(
        source,
        "positive-progression",
        semantics_path,
        output_dir,
    )

    normalized = json.loads((output_dir / "normalized-data.json").read_text())
    assert normalized["records"] == []
    assert normalized["warnings"] == [
        {
            "id": "OV-U",
            "source_row": 2,
            "code": "unknown_transition",
            "message": "Transition stage order is unknown",
            "month": "2026-04",
        }
    ]
    assert normalized["counts"]["warnings"] == 1


@pytest.mark.parametrize(
    ("name", "contents", "code", "message"),
    [
        ("malformed.json", b'{"rows": [', "invalid_json", "Source JSON is malformed"),
        (
            "invalid.csv",
            b"Area,Sub-area,Opportunity Name,TCV,Probability,Mar '26,Apr '26\n\xff",
            "invalid_encoding",
            "Source text is not valid UTF-8",
        ),
        (
            "corrupt.xlsx",
            b"not an xlsx archive",
            "invalid_xlsx",
            "Source XLSX is invalid or corrupt",
        ),
    ],
)
def test_inspect_cli_returns_safe_structured_source_errors(
    tmp_path, name, contents, code, message
):
    source = tmp_path / name
    source.write_bytes(contents)

    result = subprocess.run(
        [sys.executable, str(SCRIPTS / "prepare_opportunities.py"), "inspect", str(source)],
        check=False,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 2
    assert result.stderr == ""
    assert result.stdout.count("\n") == 1
    assert json.loads(result.stdout) == {
        "ok": False,
        "error": {
            "code": code,
            "message": message,
            "details": {"source": name},
        },
    }


def test_inspect_cli_returns_safe_structured_error_for_missing_source(tmp_path):
    source = tmp_path / "missing.json"

    result = subprocess.run(
        [sys.executable, str(SCRIPTS / "prepare_opportunities.py"), "inspect", str(source)],
        check=False,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 2
    assert result.stderr == ""
    assert result.stdout.count("\n") == 1
    assert json.loads(result.stdout) == {
        "ok": False,
        "error": {
            "code": "source_not_found",
            "message": "Source file not found",
            "details": {"source": "missing.json"},
        },
    }


def test_prepare_cli_rejects_existing_file_destination_without_overwrite(
    tmp_path, csv_source, semantics
):
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")
    output_file = tmp_path / "existing-output"
    output_file.write_text("DO NOT OVERWRITE", encoding="utf-8")

    result = subprocess.run(
        [
            sys.executable,
            str(SCRIPTS / "prepare_opportunities.py"),
            "prepare",
            str(csv_source),
            "--view",
            "wins",
            "--semantics",
            str(semantics_path),
            "--output-dir",
            str(output_file),
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 2
    assert result.stderr == ""
    assert result.stdout.count("\n") == 1
    assert json.loads(result.stdout) == {
        "ok": False,
        "error": {
            "code": "output_exists",
            "message": "Output destination already exists",
            "details": {},
        },
    }
    assert output_file.read_text(encoding="utf-8") == "DO NOT OVERWRITE"
    assert str(output_file) not in result.stdout


def test_prepare_cli_reports_atomic_write_failure_without_partial_artifacts(
    tmp_path, csv_source, semantics, monkeypatch, capsys
):
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")
    output_dir = tmp_path / "prepared"
    original_link = os.link
    link_calls = 0

    def fail_link(source, target, **kwargs):
        nonlocal link_calls
        link_calls += 1
        if link_calls == 2:
            raise OSError("sensitive raw path /private/secret")
        return original_link(source, target, **kwargs)

    monkeypatch.setattr(os, "link", fail_link)

    result = prepare_opportunities.main(
        [
            "prepare",
            str(csv_source),
            "--view",
            "wins",
            "--semantics",
            str(semantics_path),
            "--output-dir",
            str(output_dir),
        ]
    )
    captured = capsys.readouterr()

    assert result == 2
    assert captured.err == ""
    assert captured.out.count("\n") == 1
    assert json.loads(captured.out) == {
        "ok": False,
        "error": {
            "code": "output_unwritable",
            "message": "Unable to write output artifacts",
            "details": {},
        },
    }
    assert "/private/secret" not in captured.out
    assert not output_dir.exists() or list(output_dir.iterdir()) == []


def test_atomic_json_rejects_raced_temporary_symlink_without_touching_victim(tmp_path):
    output = tmp_path / "source-summary.json"
    temporary = tmp_path / ".source-summary.json.tmp"
    victim = tmp_path / "external-victim.txt"
    victim.write_text("DO NOT TOUCH\n", encoding="utf-8")
    temporary.symlink_to(victim)

    with pytest.raises(OSError):
        prepare_opportunities._atomic_json(output, {"safe": True})

    assert victim.read_text(encoding="utf-8") == "DO NOT TOUCH\n"
    assert temporary.is_symlink()
    assert not output.exists()


def test_atomic_json_cannot_overwrite_target_created_during_publish(tmp_path, monkeypatch):
    output = tmp_path / "source-summary.json"
    real_link = os.link

    def competing_publish(source, target, **kwargs):
        Path(target).write_text("competitor\n", encoding="utf-8")
        return real_link(source, target, **kwargs)

    monkeypatch.setattr(os, "link", competing_publish)

    with pytest.raises(FileExistsError):
        prepare_opportunities._atomic_json(output, {"owner": "preparer"})

    assert output.read_text(encoding="utf-8") == "competitor\n"
    assert not (tmp_path / ".source-summary.json.tmp").exists()


def test_atomic_json_rejects_stage_path_substitution_after_write_before_link(
    tmp_path, monkeypatch
):
    output_dir = tmp_path / "prepared"
    output_dir.mkdir()
    output = output_dir / "source-summary.json"
    temporary = output_dir / ".source-summary.json.tmp"
    expected = json.dumps({"safe": True}, indent=2, sort_keys=True) + "\n"
    victim = tmp_path / "external-victim.json"
    victim.write_text(expected, encoding="utf-8")
    real_link = os.link
    calls = 0

    def substitute_stage_then_link(source, target, **kwargs):
        nonlocal calls
        calls += 1
        source = Path(source)
        if calls == 1:
            source.unlink()
            source.symlink_to(victim)
        return real_link(source, target, **kwargs)

    monkeypatch.setattr(os, "link", substitute_stage_then_link)

    with pytest.raises(DataContractError) as caught:
        prepare_opportunities._write_artifacts(
            output_dir,
            [("source-summary.json", {"safe": True})],
            create_output_dir=False,
        )

    assert caught.value.code == "output_unwritable"
    victim.write_text("MUTATED EXTERNAL CONTENT\n", encoding="utf-8")
    assert output.is_symlink()
    assert output.read_text(encoding="utf-8") == "MUTATED EXTERNAL CONTENT\n"
    assert temporary.is_symlink()
    assert victim.read_text(encoding="utf-8") == "MUTATED EXTERNAL CONTENT\n"


def test_atomic_json_rejects_same_inode_mutation_before_initial_fd_hash(
    tmp_path, monkeypatch
):
    output_dir = tmp_path / "prepared"
    output_dir.mkdir()
    temporary = output_dir / ".source-summary.json.tmp"
    real_fstat = os.fstat
    calls = 0

    def mutate_then_fstat(fd):
        nonlocal calls
        calls += 1
        if calls == 1:
            temporary.write_text("CORRUPTED AFTER WRITE\n", encoding="utf-8")
        return real_fstat(fd)

    monkeypatch.setattr(os, "fstat", mutate_then_fstat)

    with pytest.raises(DataContractError) as caught:
        prepare_opportunities._write_artifacts(
            output_dir,
            [("source-summary.json", {"safe": True})],
            create_output_dir=False,
        )

    assert caught.value.code == "output_unwritable"
    assert not (output_dir / "source-summary.json").exists()
    assert not temporary.exists()


def test_preparation_rollback_preserves_competitors_for_all_artifact_names(
    tmp_path, monkeypatch
):
    output_dir = tmp_path / "prepared"
    real_link = os.link
    calls = 0
    competitor_stats = {}

    def replace_owned_outputs_then_claim_last(source, target, **kwargs):
        nonlocal calls
        calls += 1
        target = Path(target)
        if calls == 3:
            for name in ("source-summary.json", "normalized-data.json"):
                path = output_dir / name
                path.unlink()
                path.write_text(f"competitor-{name}\n", encoding="utf-8")
                competitor_stats[name] = path.lstat()
            target.write_text("competitor-exclusions.json\n", encoding="utf-8")
            competitor_stats[target.name] = target.lstat()
        return real_link(source, target, **kwargs)

    monkeypatch.setattr(os, "link", replace_owned_outputs_then_claim_last)

    with pytest.raises(DataContractError) as caught:
        prepare_opportunities._write_artifacts(
            output_dir,
            [
                ("source-summary.json", {"artifact": 1}),
                ("normalized-data.json", {"artifact": 2}),
                ("exclusions.json", {"artifact": 3}),
            ],
            create_output_dir=True,
        )

    assert caught.value.code == "output_unwritable"
    for name, expected_stat in competitor_stats.items():
        path = output_dir / name
        assert os.path.samestat(path.lstat(), expected_stat)
        assert path.read_text(encoding="utf-8") == f"competitor-{name}\n"
    assert not list(output_dir.glob(".*.tmp"))


@pytest.mark.parametrize("control", [KeyboardInterrupt(), SystemExit(23)])
def test_write_artifacts_reraises_process_control_exceptions(
    tmp_path, monkeypatch, control
):
    output_dir = tmp_path / "prepared"
    original_atomic_json = prepare_opportunities._atomic_json
    calls = 0

    def interrupt(path, value):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise control
        return original_atomic_json(path, value)

    monkeypatch.setattr(prepare_opportunities, "_atomic_json", interrupt)

    with pytest.raises(type(control)) as caught:
        prepare_opportunities._write_artifacts(
            output_dir,
            [
                ("source-summary.json", {"safe": True}),
                ("normalized-data.json", {"safe": True}),
            ],
            create_output_dir=True,
        )

    assert caught.value is control
    assert not output_dir.exists()


def test_prepare_cli_rejects_symlink_destination_without_touching_target(
    tmp_path, csv_source, semantics
):
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")
    target_dir = tmp_path / "target"
    target_dir.mkdir()
    output_link = tmp_path / "linked-output"
    output_link.symlink_to(target_dir, target_is_directory=True)

    result = subprocess.run(
        [
            sys.executable,
            str(SCRIPTS / "prepare_opportunities.py"),
            "prepare",
            str(csv_source),
            "--view",
            "wins",
            "--semantics",
            str(semantics_path),
            "--output-dir",
            str(output_link),
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 2
    assert result.stderr == ""
    assert result.stdout.count("\n") == 1
    assert json.loads(result.stdout) == {
        "ok": False,
        "error": {
            "code": "output_exists",
            "message": "Output destination already exists",
            "details": {},
        },
    }
    assert output_link.is_symlink()
    assert list(target_dir.iterdir()) == []
    assert str(output_link) not in result.stdout


def _write_showcase_csv(path: Path) -> None:
    rows = _showcase_rows()
    headers = list(rows[0])
    path.write_text(
        ",".join(headers)
        + "\n"
        + "\n".join(
            ",".join(str(row[header]) for header in headers) for row in rows
        )
        + "\n",
        encoding="utf-8",
    )


def test_analyze_groups_safe_transitions_and_marks_positive_inclusion(
    tmp_path, semantics
):
    source = tmp_path / "pipeline.csv"
    _write_showcase_csv(source)
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")
    before_bytes = source.read_bytes()
    before_names = sorted(path.name for path in tmp_path.iterdir())

    positive = analyze(source, "positive-progression", semantics_path)
    all_progression = analyze(source, "all-progression", semantics_path)

    assert positive["unresolved_transitions"] == [
        {
            "from_stage": "Discovery",
            "to_stage": "Deferred",
            "code": "unknown_transition",
            "occurrences": 1,
            "affects_inclusion": True,
            "terminal_status_resolved": False,
            "affects_truncation": True,
        }
    ]
    assert all_progression["unresolved_transitions"] == [
        {
            "from_stage": "Discovery",
            "to_stage": "Deferred",
            "code": "unknown_transition",
            "occurrences": 1,
            "affects_inclusion": True,
            "terminal_status_resolved": False,
            "affects_truncation": True,
        }
    ]
    assert positive["mixed_transitions"] == [
        {
            "from_stage": "Solution",
            "to_stage": "Proposal",
            "code": "mixed_signals",
            "occurrences": 1,
            "affects_inclusion": False,
            "terminal_status_resolved": True,
            "affects_truncation": False,
        }
    ]
    assert positive["unresolved_terminal_stages"] == [
        {
            "stage": "Deferred",
            "code": "unknown_terminal_status",
            "occurrences": 3,
            "affects_output": True,
        }
    ]
    assert positive["source"] == {
        "basename": "pipeline.csv",
        "format": "csv",
        "json_key": None,
        "sha256": positive["source"]["sha256"],
        "sheet": None,
    }
    assert len(positive["source"]["sha256"]) == 64
    assert positive["counts"] == {
        "source_rows": 12,
        "normalized_rows": 11,
        "filtered_rows": 11,
        "excluded_rows": 1,
        "warnings": 3,
        "unresolved_terminal_stage_groups": 1,
        "unresolved_terminal_stage_occurrences": 3,
        "unresolved_transition_groups": 1,
        "unresolved_transition_occurrences": 1,
        "mixed_transition_groups": 1,
        "mixed_transition_occurrences": 1,
    }
    serialized = json.dumps(positive)
    assert "Kite Discovery" not in serialized
    assert "OV-011" not in serialized
    assert source.read_bytes() == before_bytes
    assert sorted(path.name for path in tmp_path.iterdir()) == before_names


def test_analyze_rerun_resolves_unknown_with_confirmed_direction(tmp_path, semantics):
    source = tmp_path / "pipeline.csv"
    _write_showcase_csv(source)
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")
    assert analyze(source, "positive-progression", semantics_path)[
        "unresolved_transitions"
    ]

    resolved = dict(semantics)
    resolved["non_terminal_stages"] = [
        *semantics["non_terminal_stages"],
        "Deferred",
    ]
    resolved["stage_paths"] = [*semantics["stage_paths"], ["Deferred", "Discovery"]]
    semantics_path.write_text(json.dumps(resolved), encoding="utf-8")

    report = analyze(source, "positive-progression", semantics_path)
    assert report["unresolved_transitions"] == []
    assert report["unresolved_terminal_stages"] == []
    assert report["counts"]["unresolved_transition_occurrences"] == 0


def test_analyze_uses_previous_populated_stage_groups_and_filters(
    tmp_path, semantics
):
    source = tmp_path / "pipeline.csv"
    source.write_text(
        "ID,Area,Sub-area,Opportunity Name,TCV,Probability,Mar '26,Apr '26,May '26\n"
        "OV-A,North,One,Hidden Name A,Large,High,Alpha,,Beta\n"
        "OV-B,South,Two,Hidden Name B,Medium,Medium,Alpha,,Beta\n",
        encoding="utf-8",
    )
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")
    filters_path = tmp_path / "filters.json"
    filters_path.write_text(json.dumps({"areas": ["north"]}), encoding="utf-8")

    grouped = analyze(source, "positive-progression", semantics_path)
    filtered = analyze(
        source,
        "positive-progression",
        semantics_path,
        filters_path=filters_path,
    )

    assert grouped["unresolved_transitions"] == [
        {
            "from_stage": "Alpha",
            "to_stage": "Beta",
            "code": "unknown_transition",
            "occurrences": 2,
            "affects_inclusion": True,
            "terminal_status_resolved": False,
            "affects_truncation": True,
        }
    ]
    assert filtered["unresolved_transitions"][0]["occurrences"] == 1
    assert filtered["counts"]["filtered_rows"] == 1
    assert filtered["filter_keys"] == ["areas"]
    assert "Hidden Name" not in json.dumps(filtered)


def test_non_terminal_stages_are_casefolded_and_disjoint_from_terminals(semantics):
    legacy = dict(semantics)
    legacy.pop("non_terminal_stages")
    assert validate_semantics(legacy)["non_terminal_stages"] == []

    casefolded = dict(semantics)
    casefolded["non_terminal_stages"] = ["deferred"]
    validated = validate_semantics(casefolded)
    assert validated["non_terminal_stages"] == ["deferred"]

    overlapping = dict(semantics)
    overlapping["positive_terminals"] = ["Won", "Deferred"]
    overlapping["non_terminal_stages"] = ["deFERred"]
    with pytest.raises(DataContractError) as error:
        validate_semantics(overlapping)
    assert error.value.code == "invalid_semantics"
    assert str(error.value) == "Terminal and non-terminal stage lists overlap"

    duplicated = dict(semantics)
    duplicated["non_terminal_stages"] = ["Deferred", "deferred"]
    with pytest.raises(DataContractError) as error:
        validate_semantics(duplicated)
    assert error.value.code == "invalid_semantics"

    blank = dict(semantics)
    blank["non_terminal_stages"] = ["   "]
    with pytest.raises(DataContractError) as error:
        validate_semantics(blank)
    assert error.value.code == "invalid_semantics"


@pytest.mark.parametrize(
    "view", ["wins", "losses", "all-progression", "positive-progression"]
)
def test_analyze_requires_terminal_confirmation_for_unknown_positive_alias(
    tmp_path, semantics, view
):
    source = tmp_path / "positive-alias.csv"
    source.write_text(
        "ID,Area,Sub-area,Opportunity Name,TCV,Probability,Mar '26,Apr '26,May '26\n"
        "OV-A,North,One,Secret Alias,Large,High,Proposal,Closed Won,Proposal\n",
        encoding="utf-8",
    )
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")
    before = source.read_bytes()

    report = analyze(source, view, semantics_path)

    assert report["unresolved_terminal_stages"] == [
        {
            "stage": "Closed Won",
            "code": "unknown_terminal_status",
            "occurrences": 1,
            "affects_output": True,
        }
    ]
    transition = next(
        item
        for item in report["unresolved_transitions"]
        if item["from_stage"] == "Proposal" and item["to_stage"] == "Closed Won"
    )
    assert transition["terminal_status_resolved"] is False
    assert transition["affects_inclusion"] is True
    assert transition["affects_truncation"] is True
    assert "Secret Alias" not in json.dumps(report)
    assert "OV-A" not in json.dumps(report)
    assert source.read_bytes() == before
    assert sorted(path.name for path in tmp_path.iterdir()) == [
        "positive-alias.csv",
        "semantics.json",
    ]


def test_confirmed_positive_alias_enables_wins_and_terminal_cutoff(
    tmp_path, semantics
):
    source = tmp_path / "positive-alias.csv"
    source.write_text(
        "ID,Area,Sub-area,Opportunity Name,TCV,Probability,Mar '26,Apr '26,May '26\n"
        "OV-A,North,One,Secret Alias,Large,High,Proposal,Closed Won,Proposal\n",
        encoding="utf-8",
    )
    confirmed = dict(semantics)
    confirmed["positive_terminals"] = [*semantics["positive_terminals"], "closed won"]
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(confirmed), encoding="utf-8")

    report = analyze(source, "wins", semantics_path)
    assert report["unresolved_terminal_stages"] == []
    output_dir = tmp_path / "wins"
    prepare(source, "wins", semantics_path, output_dir)
    normalized = json.loads((output_dir / "normalized-data.json").read_text())
    assert [record["id"] for record in normalized["records"]] == ["OV-A"]
    assert normalized["records"][0]["terminal"] == {
        "kind": "positive",
        "month": "2026-04",
        "index": 1,
    }
    assert [month["stage"] for month in normalized["records"][0]["months"]] == [
        "Proposal",
        "Closed Won",
    ]


@pytest.mark.parametrize("view", ["all-progression", "positive-progression"])
def test_unknown_negative_alias_flags_terminal_impact_then_truncates_when_confirmed(
    tmp_path, semantics, view
):
    source = tmp_path / "negative-alias.csv"
    source.write_text(
        "ID,Area,Sub-area,Opportunity Name,TCV,Probability,Mar '26,Apr '26,May '26\n"
        "OV-N,North,One,Secret Loss,Large,High,Proposal,Closed Lost,Won\n",
        encoding="utf-8",
    )
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")

    unresolved = analyze(source, view, semantics_path)
    transition = next(
        item
        for item in unresolved["unresolved_transitions"]
        if item["to_stage"] == "Closed Lost"
    )
    assert transition["terminal_status_resolved"] is False
    assert transition["affects_inclusion"] is True
    assert transition["affects_truncation"] is True

    confirmed = dict(semantics)
    confirmed["negative_terminals"] = [
        *semantics["negative_terminals"],
        "closed lost",
    ]
    semantics_path.write_text(json.dumps(confirmed), encoding="utf-8")
    assert analyze(source, view, semantics_path)["unresolved_terminal_stages"] == []
    output_dir = tmp_path / view
    prepare(source, view, semantics_path, output_dir)
    normalized = json.loads((output_dir / "normalized-data.json").read_text())
    if view == "all-progression":
        assert [record["id"] for record in normalized["records"]] == ["OV-N"]
        assert normalized["records"][0]["terminal"]["kind"] == "negative"
        assert [month["stage"] for month in normalized["records"][0]["months"]] == [
            "Proposal",
            "Closed Lost",
        ]
    else:
        assert normalized["records"] == []


def test_analyze_cli_prints_one_json_object_without_artifacts(tmp_path, semantics):
    source = tmp_path / "pipeline.csv"
    _write_showcase_csv(source)
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")
    before = sorted(path.name for path in tmp_path.iterdir())

    result = subprocess.run(
        [
            sys.executable,
            str(SCRIPTS / "prepare_opportunities.py"),
            "analyze",
            str(source),
            "--view",
            "positive-progression",
            "--semantics",
            str(semantics_path),
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 0
    assert result.stderr == ""
    assert result.stdout.count("\n") == 1
    payload = json.loads(result.stdout)
    assert payload["ok"] is True
    assert payload["unresolved_transitions"][0]["from_stage"] == "Discovery"
    assert sorted(path.name for path in tmp_path.iterdir()) == before


def test_analyze_cli_returns_safe_structured_error_without_traceback(
    tmp_path, semantics
):
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")

    result = subprocess.run(
        [
            sys.executable,
            str(SCRIPTS / "prepare_opportunities.py"),
            "analyze",
            str(tmp_path / "missing.csv"),
            "--view",
            "positive-progression",
            "--semantics",
            str(semantics_path),
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 2
    assert result.stderr == ""
    assert result.stdout.count("\n") == 1
    assert json.loads(result.stdout)["error"] == {
        "code": "source_not_found",
        "message": "Source file not found",
        "details": {"source": "missing.csv"},
    }


def test_analyze_cli_returns_structured_json_for_invalid_arguments(tmp_path):
    source = tmp_path / "pipeline.csv"
    source.write_text("Area\nNorth\n", encoding="utf-8")

    result = subprocess.run(
        [
            sys.executable,
            str(SCRIPTS / "prepare_opportunities.py"),
            "analyze",
            str(source),
            "--view",
            "positive-progression",
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 2
    assert result.stderr == ""
    assert result.stdout.count("\n") == 1
    assert json.loads(result.stdout)["error"] == {
        "code": "invalid_arguments",
        "message": "Invalid command arguments",
        "details": {},
    }


@pytest.mark.parametrize(
    ("alias", "terminal_key", "later_terminal"),
    [
        ("Closed Lost", "negative_terminals", "Won"),
        ("Closed Won", "positive_terminals", "Lost"),
    ],
)
@pytest.mark.parametrize(
    "view", ["wins", "losses", "all-progression", "positive-progression"]
)
def test_analyze_reports_pre_range_terminal_alias_before_stable_exclusion(
    tmp_path, semantics, alias, terminal_key, later_terminal, view
):
    source = tmp_path / "pre-range-alias.csv"
    source.write_text(
        "ID,Area,Sub-area,Opportunity Name,TCV,Probability,Mar '26,Apr '26,May '26\n"
        f"OV-A,North,One,Secret Alias,Large,High,{alias},Proposal,{later_terminal}\n"
        f"OV-B,North,One,Blank Before,Medium,Medium,,Proposal,{later_terminal}\n"
        f"OV-C,North,One,Known Before,Small,Low,Discovery,Proposal,{later_terminal}\n",
        encoding="utf-8",
    )
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")
    before_bytes = source.read_bytes()
    before_names = sorted(path.name for path in tmp_path.iterdir())

    report = analyze(
        source,
        view,
        semantics_path,
        months=["Apr '26", "May '26"],
    )

    assert report["unresolved_terminal_stages"] == [
        {
            "stage": alias,
            "code": "unknown_terminal_status",
            "occurrences": 1,
            "affects_output": True,
        }
    ]
    assert report["unresolved_transitions"] == []
    serialized = json.dumps(report)
    for private_value in ("OV-A", "Secret Alias", "Blank Before", "Known Before"):
        assert private_value not in serialized
    assert source.read_bytes() == before_bytes
    assert sorted(path.name for path in tmp_path.iterdir()) == before_names

    confirmed = dict(semantics)
    confirmed[terminal_key] = [*semantics[terminal_key], alias.casefold()]
    semantics_path.write_text(json.dumps(confirmed), encoding="utf-8")
    output_dir = tmp_path / view
    prepare(
        source,
        view,
        semantics_path,
        output_dir,
        months=["Apr '26", "May '26"],
    )
    normalized = json.loads((output_dir / "normalized-data.json").read_text())
    assert "OV-A" not in [record["id"] for record in normalized["records"]]
    assert next(
        item for item in normalized["exclusions"] if item["id"] == "OV-A"
    )["code"] == "terminal_before_range"


def test_analyze_rejects_uncached_formula_before_pre_range_terminal(
    tmp_path, semantics
):
    source = tmp_path / "pre-range-formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pipeline"
    headers = [*HEADERS, "Jun '26"]
    sheet.append(headers)
    sheet.append(
        [
            "North",
            "Cloud",
            "Project Cedar",
            "Large",
            "High",
            '=IF(1=1,"Closed Lost","")',
            "Lost",
            "Proposal",
            "Won",
        ]
    )
    workbook.save(source)
    workbook.close()
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")

    with pytest.raises(DataContractError) as error:
        analyze(
            source,
            "wins",
            semantics_path,
            sheet="Pipeline",
            months=["May '26", "Jun '26"],
        )

    assert error.value.code == "formula_cache_missing"
    assert error.value.details == {"cells": [{"row": 2, "header": "Mar '26"}]}


@pytest.mark.parametrize(("terminal", "later_terminal"), [("Lost", "Won"), ("Won", "Lost")])
@pytest.mark.parametrize(
    "view", ["wins", "losses", "all-progression", "positive-progression"]
)
def test_analyze_stops_pre_range_scan_at_first_confirmed_terminal(
    tmp_path, semantics, terminal, later_terminal, view
):
    source = tmp_path / "first-terminal.csv"
    source.write_text(
        "ID,Area,Sub-area,Opportunity Name,TCV,Probability,Mar '26,Apr '26,May '26,Jun '26\n"
        f"OV-T,North,One,Private Terminal,Large,High,{terminal},Deferred,Proposal,{later_terminal}\n",
        encoding="utf-8",
    )
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")
    before_bytes = source.read_bytes()
    before_names = sorted(path.name for path in tmp_path.iterdir())

    report = analyze(
        source,
        view,
        semantics_path,
        months=["May '26", "Jun '26"],
    )

    assert report["unresolved_terminal_stages"] == []
    assert report["unresolved_transitions"] == []
    serialized = json.dumps(report)
    assert "Deferred" not in serialized
    assert "Private Terminal" not in serialized
    assert "OV-T" not in serialized
    assert source.read_bytes() == before_bytes
    assert sorted(path.name for path in tmp_path.iterdir()) == before_names

    output_dir = tmp_path / view
    prepare(
        source,
        view,
        semantics_path,
        output_dir,
        months=["May '26", "Jun '26"],
    )
    normalized = json.loads((output_dir / "normalized-data.json").read_text())
    assert normalized["records"] == []
    assert normalized["exclusions"] == [
        {
            "id": "OV-T",
            "source_row": 2,
            "code": "terminal_before_range",
            "message": "Row reached a terminal before the selected range",
        }
    ]


def test_analyze_keeps_unresolved_pre_range_stage_before_terminal(tmp_path, semantics):
    source = tmp_path / "unknown-before-terminal.csv"
    source.write_text(
        "ID,Area,Sub-area,Opportunity Name,TCV,Probability,Mar '26,Apr '26,May '26,Jun '26\n"
        "OV-U,North,One,Private Unknown,Large,High,Deferred,Lost,Proposal,Won\n",
        encoding="utf-8",
    )
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")

    report = analyze(
        source,
        "wins",
        semantics_path,
        months=["May '26", "Jun '26"],
    )

    assert report["unresolved_terminal_stages"] == [
        {
            "stage": "Deferred",
            "code": "unknown_terminal_status",
            "occurrences": 1,
            "affects_output": True,
        }
    ]
    assert report["unresolved_transitions"] == []


def test_analyze_ignores_uncached_pre_range_formula_after_terminal(tmp_path, semantics):
    source = tmp_path / "formula-after-terminal.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pipeline"
    sheet.append([*HEADERS, "Jun '26"])
    sheet.append(
        [
            "North",
            "Cloud",
            "Private Formula",
            "Large",
            "High",
            "Lost",
            '=IF(1=1,"Deferred","")',
            "Proposal",
            "Won",
        ]
    )
    workbook.save(source)
    workbook.close()
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")

    report = analyze(
        source,
        "wins",
        semantics_path,
        sheet="Pipeline",
        months=["May '26", "Jun '26"],
    )

    assert report["unresolved_terminal_stages"] == []
    assert "Deferred" not in json.dumps(report)


def test_selected_formula_after_pre_range_terminal_does_not_block(
    tmp_path, semantics
):
    source = tmp_path / "selected-formula-after-terminal.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pipeline"
    sheet.append([*HEADERS, "Jun '26"])
    sheet.append(
        [
            "North",
            "Cloud",
            "Private Selected Formula",
            "Large",
            "High",
            "Lost",
            "Deferred",
            '=IF(1=1,"Proposal","")',
            "Won",
        ]
    )
    workbook.save(source)
    workbook.close()
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")

    report = analyze(
        source,
        "wins",
        semantics_path,
        sheet="Pipeline",
        months=["May '26", "Jun '26"],
    )
    assert report["unresolved_terminal_stages"] == []

    output_dir = tmp_path / "wins"
    prepare(
        source,
        "wins",
        semantics_path,
        output_dir,
        sheet="Pipeline",
        months=["May '26", "Jun '26"],
    )
    normalized = json.loads((output_dir / "normalized-data.json").read_text())
    assert normalized["records"] == []
    assert normalized["exclusions"][0]["code"] == "terminal_before_range"


def test_formula_before_terminal_suppresses_only_later_row_formulas(
    tmp_path, semantics
):
    source = tmp_path / "formula-before-terminal.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pipeline"
    sheet.append([*HEADERS, "Jun '26"])
    sheet.append(
        [
            "North",
            "Cloud",
            "Private Formula Order",
            "Large",
            "High",
            '=IF(1=1,"Deferred","")',
            "Lost",
            '=IF(1=1,"Proposal","")',
            "Won",
        ]
    )
    workbook.save(source)
    workbook.close()
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")

    with pytest.raises(DataContractError) as error:
        analyze(
            source,
            "wins",
            semantics_path,
            sheet="Pipeline",
            months=["May '26", "Jun '26"],
        )

    assert error.value.code == "formula_cache_missing"
    assert error.value.details == {"cells": [{"row": 2, "header": "Mar '26"}]}


def test_fixed_formulas_on_pre_range_excluded_row_do_not_block(tmp_path, semantics):
    source = tmp_path / "fixed-formulas-after-terminal.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pipeline"
    sheet.append([*HEADERS, "Jun '26"])
    sheet.append(
        [
            "North",
            "Cloud",
            '="Private Fixed Formula"',
            '=1+1',
            "High",
            "Lost",
            "Deferred",
            "Proposal",
            "Won",
        ]
    )
    workbook.save(source)
    workbook.close()
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")

    report = analyze(
        source,
        "wins",
        semantics_path,
        sheet="Pipeline",
        months=["May '26", "Jun '26"],
    )
    assert report["unresolved_terminal_stages"] == []

    output_dir = tmp_path / "wins"
    prepare(
        source,
        "wins",
        semantics_path,
        output_dir,
        sheet="Pipeline",
        months=["May '26", "Jun '26"],
    )
    normalized = json.loads((output_dir / "normalized-data.json").read_text())
    assert normalized["records"] == []
    assert normalized["exclusions"] == [
        {
            "id": "ROW-0002",
            "source_row": 2,
            "code": "terminal_before_range",
            "message": "Row reached a terminal before the selected range",
        }
    ]


def test_formula_without_later_terminal_keeps_fixed_and_selected_errors(
    tmp_path, semantics
):
    source = tmp_path / "formula-without-terminal.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pipeline"
    sheet.append([*HEADERS, "Jun '26"])
    sheet.append(
        [
            "North",
            "Cloud",
            '="Private Active Formula"',
            '=1+1',
            "High",
            '=IF(1=1,"Deferred","")',
            "Discovery",
            '=IF(1=1,"Proposal","")',
            "Won",
        ]
    )
    workbook.save(source)
    workbook.close()
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")

    with pytest.raises(DataContractError) as error:
        analyze(
            source,
            "wins",
            semantics_path,
            sheet="Pipeline",
            months=["May '26", "Jun '26"],
        )

    assert error.value.code == "formula_cache_missing"
    assert {(cell["row"], cell["header"]) for cell in error.value.details["cells"]} == {
        (2, "Opportunity Name"),
        (2, "TCV"),
        (2, "Mar '26"),
        (2, "May '26"),
    }


@pytest.mark.parametrize(
    "view", ["wins", "losses", "all-progression", "positive-progression"]
)
def test_post_formula_control_scan_hides_intervening_unknown_stage(
    tmp_path, semantics, view
):
    source = tmp_path / "private-control-scan.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pipeline"
    sheet.append([*HEADERS, "Jun '26", "Jul '26"])
    sheet.append(
        [
            "North",
            "Cloud",
            "Private Control Scan",
            "Large",
            "High",
            '=IF(1=1,"Maybe Terminal","")',
            "Secret Deferred",
            "Lost",
            "Proposal",
            "Won",
        ]
    )
    workbook.save(source)
    workbook.close()
    semantics_path = tmp_path / "semantics.json"
    semantics_path.write_text(json.dumps(semantics), encoding="utf-8")
    before_bytes = source.read_bytes()
    before_names = sorted(path.name for path in tmp_path.iterdir())

    with pytest.raises(DataContractError) as error:
        analyze(
            source,
            view,
            semantics_path,
            sheet="Pipeline",
            months=["Jun '26", "Jul '26"],
        )

    assert error.value.details == {"cells": [{"row": 2, "header": "Mar '26"}]}
    serialized = json.dumps(error.value.details)
    for private_value in ("Secret Deferred", "Private Control Scan", "Maybe Terminal"):
        assert private_value not in serialized
    assert source.read_bytes() == before_bytes
    assert sorted(path.name for path in tmp_path.iterdir()) == before_names
