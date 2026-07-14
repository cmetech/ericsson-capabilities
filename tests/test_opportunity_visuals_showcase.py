import hashlib
import json
import re
import subprocess
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree

import pytest
from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[1]
SKILL = REPO / "skills/ericsson/opportunity-visuals"
SCRIPTS = SKILL / "scripts"
FIXTURES = REPO / "tests/fixtures/opportunity_visuals"
GOLDEN = REPO / "tests/golden/opportunity_visuals"
SHOWCASE_CSV = FIXTURES / "showcase-opportunities.csv"
SHOWCASE_JSON = FIXTURES / "showcase-opportunities.json"
SHOWCASE_XLSX = FIXTURES / "showcase-opportunities.xlsx"

sys.path.insert(0, str(SCRIPTS))

from prepare_opportunities import analyze, prepare  # noqa: E402
from render_opportunity_visual import render_document  # noqa: E402


REQUIRED_FIXTURES = (
    "build_showcase_fixtures.py",
    "showcase-opportunities.csv",
    "showcase-opportunities.json",
    "showcase-opportunities.xlsx",
    "stage-semantics.json",
    "expected-normalized.json",
    "expected-wins.json",
    "expected-losses.json",
    "expected-all-progression.json",
    "expected-positive-progression.json",
    "expected-run-summary.json",
)
REQUIRED_GOLDENS = (
    "wins-p01.svg",
    "losses-p01.svg",
    "all-progression-p01.svg",
    "positive-progression-p01.svg",
)
EXPECTED_NAMES = {
    "Aurora Core Renewal",
    "Beacon Automation",
    "Cedar Assurance",
    "Delta Capacity",
    "Echo Modernization",
    "Fjord Analytics",
    "Grove Orchestration",
    "Harbor Observability <Pilot> =1+1",
    "Ion Edge Program",
    "Juniper Expansion",
    "Kite Discovery",
    "Lumen Platform",
}
FORBIDDEN_TOKENS = (
    "customer",
    "account name",
    "verizon",
    "at&t",
    "t-mobile",
    "vodafone",
    "telefonica",
    "deutsche telekom",
    "orange business",
    "telia",
)
EXPECTED_WARNINGS = [
    {"id": "OV-006", "code": "skipped_blank_months"},
    {"id": "OV-007", "code": "mixed_signals"},
    {"id": "OV-011", "code": "unknown_transition"},
]
APPROVED_SVG_PALETTE = {
    "#000000",
    "#1174E6",
    "#23969A",
    "#A6A6A6",
    "#D8D8D8",
    "#E65D6A",
    "#F2F2F2",
    "#FFFFFF",
}
SVG_HEX_COLOR = re.compile(
    r"(?<![0-9a-f])#(?:[0-9a-f]{8}|[0-9a-f]{6}|[0-9a-f]{4}|[0-9a-f]{3})(?![0-9a-f])",
    re.IGNORECASE,
)


def load_json(path: Path) -> dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))


def project_view(output_dir: Path) -> dict[str, object]:
    document = load_json(output_dir / "normalized-data.json")
    exclusions = load_json(output_dir / "exclusions.json")["exclusions"]
    return {
        "included_ids": [record["id"] for record in document["records"]],
        "excluded": [{"id": item["id"], "code": item["code"]} for item in exclusions],
        "transitions": {
            record["id"]: [
                month["classification"] if month["stage"] else "empty"
                for month in record["months"]
            ]
            for record in document["records"]
        },
        "terminals": {
            record["id"]: record["terminal"] for record in document["records"]
        },
    }


def canonical_svg(path: Path) -> str:
    root = ElementTree.parse(path).getroot()
    return ElementTree.tostring(root, encoding="unicode")


def svg_hex_colors(root: ElementTree.Element) -> set[str]:
    values = [value for element in root.iter() for value in element.attrib.values()]
    values.extend(
        element.text or ""
        for element in root.iter()
        if element.tag.rsplit("}", 1)[-1].casefold() == "style"
    )
    return {
        match.group(0).upper()
        for value in values
        for match in SVG_HEX_COLOR.finditer(value)
    }


def assert_approved_svg_palette(root: ElementTree.Element) -> set[str]:
    colors = svg_hex_colors(root)
    unapproved = colors - APPROVED_SVG_PALETTE
    assert not unapproved, f"unapproved SVG colors: {sorted(unapproved)}"
    return colors


def prepare_source(source: Path, output_dir: Path) -> dict[str, object]:
    options = {"sheet": "Pipeline"} if source.suffix == ".xlsx" else {}
    prepare(
        source,
        "all-progression",
        FIXTURES / "stage-semantics.json",
        output_dir,
        **options,
    )
    return load_json(output_dir / "normalized-data.json")


def project_source_format(document: dict[str, object]) -> dict[str, object]:
    projected = dict(document)
    projected.pop("source")
    return projected


def test_showcase_pack_contains_every_committed_artifact():
    missing = [
        str(path)
        for path in (
            *(FIXTURES / name for name in REQUIRED_FIXTURES),
            *(GOLDEN / name for name in REQUIRED_GOLDENS),
        )
        if not path.is_file()
    ]
    assert missing == []


def test_showcase_semantics_leave_only_deferred_terminal_status_unresolved():
    semantics = load_json(FIXTURES / "stage-semantics.json")
    assert semantics["non_terminal_stages"] == ["Discovery"]

    report = analyze(
        SHOWCASE_CSV,
        "positive-progression",
        FIXTURES / "stage-semantics.json",
    )
    assert report["unresolved_terminal_stages"] == [
        {
            "stage": "Deferred",
            "code": "unknown_terminal_status",
            "occurrences": 3,
            "affects_output": True,
        }
    ]
    assert report["unresolved_transitions"] == [
        {
            "from_stage": "Discovery",
            "to_stage": "Deferred",
            "code": "unknown_transition",
            "occurrences": 1,
            "affects_inclusion": True,
            "terminal_status_resolved": False,
            "affects_truncation": True,
        }
    ]


def test_showcase_builder_is_byte_deterministic(tmp_path):
    builder = FIXTURES / "build_showcase_fixtures.py"
    subprocess.run([sys.executable, str(builder), "--output-dir", str(tmp_path)], check=True)
    first = {
        path.name: hashlib.sha256(path.read_bytes()).hexdigest()
        for path in tmp_path.iterdir()
    }
    subprocess.run([sys.executable, str(builder), "--output-dir", str(tmp_path)], check=True)
    second = {
        path.name: hashlib.sha256(path.read_bytes()).hexdigest()
        for path in tmp_path.iterdir()
    }
    assert first == second
    assert set(first) == set(REQUIRED_FIXTURES) - {"build_showcase_fixtures.py"}
    assert second == {
        name: hashlib.sha256((FIXTURES / name).read_bytes()).hexdigest()
        for name in second
    }


def test_showcase_formats_normalize_equivalently_with_exact_provenance(tmp_path):
    documents = {
        source.suffix: prepare_source(source, tmp_path / source.suffix.removeprefix("."))
        for source in (SHOWCASE_CSV, SHOWCASE_JSON, SHOWCASE_XLSX)
    }
    assert project_source_format(documents[".csv"]) == project_source_format(
        documents[".json"]
    ) == project_source_format(documents[".xlsx"])

    for suffix, source in (
        (".csv", SHOWCASE_CSV),
        (".json", SHOWCASE_JSON),
        (".xlsx", SHOWCASE_XLSX),
    ):
        assert documents[suffix]["source"] == {
            "basename": source.name,
            "sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
            "sheet": "Pipeline" if suffix == ".xlsx" else None,
        }


def test_showcase_expected_normalization_is_independent_and_exact(tmp_path):
    document = prepare_source(SHOWCASE_CSV, tmp_path / "normalized")
    expected = load_json(FIXTURES / "expected-normalized.json")
    expected_summary = load_json(FIXTURES / "expected-run-summary.json")
    actual_records = {record["id"]: record for record in document["records"]}

    assert expected["source_ids"] == [f"OV-{number:03d}" for number in range(1, 13)]
    assert [record["id"] for record in document["records"]] == expected["included_ids"]
    assert {
        row_id: [month["classification"] for month in record["months"]]
        for row_id, record in actual_records.items()
    } == {
        row_id: transitions
        for row_id, transitions in expected["transitions"].items()
        if row_id != "OV-005"
    }
    assert expected["transitions"]["OV-005"] == ["initial"]
    assert [
        {"id": item["id"], "code": item["code"]}
        for item in document["exclusions"]
    ] == expected["excluded"]
    assert {
        row_id: record["terminal"] for row_id, record in actual_records.items()
    } == expected["terminals"]
    assert actual_records["OV-006"]["months"][1] == expected["blank_month"]
    assert actual_records["OV-006"]["months"][2]["skipped_months"] == ["Apr '26"]
    assert len(actual_records["OV-001"]["months"]) == 3
    assert len(actual_records["OV-009"]["months"]) == 2
    assert [
        {"id": item["id"], "code": item["code"]} for item in document["warnings"]
    ] == expected_summary["warnings"] == EXPECTED_WARNINGS


@pytest.mark.parametrize(
    ("view", "expected_name", "golden_name"),
    [
        ("wins", "expected-wins.json", "wins-p01.svg"),
        ("losses", "expected-losses.json", "losses-p01.svg"),
        (
            "all-progression",
            "expected-all-progression.json",
            "all-progression-p01.svg",
        ),
        (
            "positive-progression",
            "expected-positive-progression.json",
            "positive-progression-p01.svg",
        ),
    ],
)
def test_showcase_view_end_to_end(tmp_path, view, expected_name, golden_name):
    prepared = tmp_path / view
    prepare(
        SHOWCASE_CSV,
        view,
        FIXTURES / "stage-semantics.json",
        prepared,
    )
    expected = load_json(FIXTURES / expected_name)
    expected_summary = load_json(FIXTURES / "expected-run-summary.json")["views"][view]
    assert set(expected) == {"included_ids", "excluded", "transitions", "terminals"}
    assert project_view(prepared) == expected
    document = load_json(prepared / "normalized-data.json")
    assert [
        {"id": item["id"], "code": item["code"]} for item in document["warnings"]
    ] == load_json(FIXTURES / "expected-run-summary.json")["warnings"]
    assert {
        "included_rows": document["counts"]["included_rows"],
        "excluded_rows": document["counts"]["excluded_rows"],
        "warnings": document["counts"]["warnings"],
    } == {
        key: expected_summary[key]
        for key in ("included_rows", "excluded_rows", "warnings")
    }

    rendered = render_document(
        prepared / "normalized-data.json",
        prepared,
        png_mode="never",
    )
    assert len(rendered["pages"]) == 1
    assert canonical_svg(Path(rendered["pages"][0]["svg"])) == canonical_svg(
        GOLDEN / golden_name
    )
    assert rendered["png"]["status"] == "disabled"
    manifest = load_json(Path(rendered["manifest"]))
    assert len(manifest["pages"]) == expected_summary["pages"]
    assert [
        {"number": page["number"], "row_ids": page["row_ids"]}
        for page in manifest["pages"]
    ] == expected_summary["page_assignments"]

    html = Path(rendered["pages"][0]["html"]).read_text(encoding="utf-8")
    assert "https://" not in html.casefold()
    assert html.casefold().count("http://") == 1
    assert 'href="http' not in html.casefold()
    assert 'src="http' not in html.casefold()


def test_showcase_sources_are_synthetic_and_privacy_safe():
    rows = json.loads(SHOWCASE_JSON.read_text(encoding="utf-8"))
    assert len(rows) == 12
    assert {row["Opportunity Name"] for row in rows} == EXPECTED_NAMES
    assert [row["ID"] for row in rows] == [f"OV-{number:03d}" for number in range(1, 13)]
    corpus = json.dumps(rows, ensure_ascii=False).casefold()
    assert all(token not in corpus for token in FORBIDDEN_TOKENS)
    single_stage = next(row for row in rows if row["ID"] == "OV-005")
    assert [
        single_stage[label]
        for label in ("Mar '26", "Apr '26", "May '26", "Jun '26")
    ] == ["", "", "Discovery", ""]
    assert next(row for row in rows if row["ID"] == "OV-001")["Jun '26"] == (
        "In Delivery"
    )
    restarted = next(row for row in rows if row["ID"] == "OV-009")
    assert [restarted["May '26"], restarted["Jun '26"]] == ["Restarted", "Won"]


def test_showcase_xlsx_is_one_visible_plain_value_sheet():
    workbook = load_workbook(SHOWCASE_XLSX, data_only=False, read_only=False)
    try:
        assert workbook.sheetnames == ["Pipeline"]
        assert workbook["Pipeline"].sheet_state == "visible"
        assert workbook.vba_archive is None
        assert workbook._external_links == []
        assert all(
            cell.data_type != "f"
            for row in workbook["Pipeline"].iter_rows()
            for cell in row
        )
    finally:
        workbook.close()

    with zipfile.ZipFile(SHOWCASE_XLSX) as archive:
        lowered = [name.casefold() for name in archive.namelist()]
    assert not any("externallinks" in name for name in lowered)
    assert not any("vbaproject" in name or name.endswith(".bin") for name in lowered)


def test_golden_svgs_follow_the_reviewed_local_visual_contract():
    forbidden_elements = {"a", "foreignObject", "iframe", "image", "script", "use"}
    for name in REQUIRED_GOLDENS:
        path = GOLDEN / name
        text = path.read_text(encoding="utf-8")
        root = ElementTree.parse(path).getroot()
        colors = assert_approved_svg_palette(root)
        assert root.attrib["width"] == "1920"
        assert root.attrib["height"] == "1080"
        assert root.attrib["viewBox"] == "0 0 1920 1080"
        assert {
            "#000000",
            "#1174E6",
            "#A6A6A6",
            "#D8D8D8",
            "#E65D6A",
            "#F2F2F2",
            "#FFFFFF",
        } <= colors
        if name != "losses-p01.svg":
            assert "#23969A" in colors
        assert text.casefold().count("http://") == 1
        assert "https://" not in text.casefold()
        assert all(
            element.tag.rsplit("}", 1)[-1] not in forbidden_elements
            for element in root.iter()
        )
        assert "In Delivery" not in text
        assert "Restarted" not in text

    all_progression = (GOLDEN / "all-progression-p01.svg").read_text(
        encoding="utf-8"
    )
    assert (
        'data-opportunity-id="OV-006" data-month="2026-04" data-empty="true"'
        in all_progression
    )
    assert 'data-full-value="Harbor Observability &lt;Pilot&gt; =1+1"' in all_progression


@pytest.mark.parametrize("location", ["css", "stroke"])
@pytest.mark.parametrize("color", ["#123", "#1234", "#123456", "#12345678"])
def test_svg_palette_validation_rejects_every_unapproved_hex_form(location, color):
    payload = (
        f"<style>.unapproved{{fill:{color}}}</style>"
        if location == "css"
        else f'<line stroke="{color}" />'
    )
    root = ElementTree.fromstring(
        f'<svg xmlns="http://www.w3.org/2000/svg">{payload}</svg>'
    )

    with pytest.raises(AssertionError, match=color):
        assert_approved_svg_palette(root)


@pytest.mark.parametrize("color", ["#abc", "#abcd", "#abcdef", "#abcdef12"])
def test_svg_hex_color_extraction_returns_each_complete_token_once(color):
    root = ElementTree.fromstring(
        '<svg xmlns="http://www.w3.org/2000/svg">'
        f"<style>.sample{{fill:{color}}}</style>"
        "</svg>"
    )

    assert svg_hex_colors(root) == {color.upper()}
