"""Read-only opportunity source inspection and field discovery."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
from numbers import Real
from pathlib import Path


class DataContractError(ValueError):
    def __init__(
        self,
        code: str,
        message: str,
        details: dict[str, object] | None = None,
    ):
        super().__init__(message)
        self.code = code
        self.details = details or {}


FIELD_ALIASES = {
    "area": {"area", "sales area"},
    "sub_area": {"sub area", "sub-area", "subarea"},
    "opportunity_name": {"opportunity name", "opportunity", "deal name"},
    "tcv": {"tcv", "total contract value"},
    "probability": {"probability", "current probability"},
}
MONTHS = {
    name.lower(): index
    for index, names in enumerate(
        (
            ("jan", "january"),
            ("feb", "february"),
            ("mar", "march"),
            ("apr", "april"),
            ("may",),
            ("jun", "june"),
            ("jul", "july"),
            ("aug", "august"),
            ("sep", "sept", "september"),
            ("oct", "october"),
            ("nov", "november"),
            ("dec", "december"),
        ),
        start=1,
    )
    for name in names
}

_FIXED_FIELDS = tuple(FIELD_ALIASES)
_MONTH_WORD_PATTERN = re.compile(r"^([a-z]+)\s*'?\s*(\d{2}|\d{4})$", re.IGNORECASE)
_ISO_MONTH_PATTERN = re.compile(r"^(\d{4})-(\d{2})$")
_NUMERIC_VALUE_PATTERN = re.compile(
    r"^(?:[$€£¥])?([+-]?)(\d+(?:\.\d+)?|\.\d+)([kmb])?(%)?$",
    re.IGNORECASE,
)

DEFAULT_SEMANTICS = {
    "positive_terminals": ["Won"],
    "negative_terminals": ["Lost", "Cancelled"],
    "non_terminal_stages": [],
    "stage_paths": [],
    "positive_transitions": [["Proposal", "Workshop"]],
    "tcv_order": ["X-Small", "Small", "Medium", "Large", "X-Large"],
    "probability_order": ["Low", "Medium", "High", "Certain"],
}

_FILTER_KEYS = {
    "areas",
    "sub_areas",
    "opportunity_contains",
    "tcv_min",
    "tcv_max",
    "probability_min",
    "probability_max",
}


def _normalize_header(header: str) -> str:
    return " ".join(header.replace("’", "'").replace("_", " ").split()).casefold()


def _source_metadata(path: Path, row_count: int) -> dict[str, object]:
    return {
        "source": path.name,
        "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "row_count": row_count,
    }


def _validate_table(
    headers: list[str],
    rows: list[dict[str, object]],
    *,
    malformed_rows: list[int] | None = None,
    first_row_number: int = 2,
) -> None:
    if not headers or not rows:
        raise DataContractError("empty_source", "source contains no rows")
    normalized = [_normalize_header(header) for header in headers]
    if any(not header for header in normalized):
        raise DataContractError("blank_header", "source contains a blank header")
    duplicates = sorted({header for header in normalized if normalized.count(header) > 1})
    if duplicates:
        raise DataContractError(
            "duplicate_header",
            "source contains a duplicate header",
            {"headers": duplicates},
        )

    expected = set(headers)
    malformed = set(malformed_rows or [])
    for row_number, row in enumerate(rows, start=first_row_number):
        if row_number in malformed or set(row) != expected:
            raise DataContractError(
                "invalid_row_shape",
                f"row {row_number} does not match headers",
                {"row": row_number},
            )


def _load_source(
    path: Path,
    sheet: str | None = None,
    json_key: str | None = None,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    """Load a supported source without evaluating formulas or mutating the file."""

    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            raw_headers = reader.fieldnames or []
            headers = [header if header is not None else "" for header in raw_headers]
            raw_rows = list(reader)

        malformed_rows = [
            row_number
            for row_number, row in enumerate(raw_rows, start=2)
            if None in row or any(value is None for value in row.values())
        ]
        rows = [dict(row) for row in raw_rows]
        _validate_table(headers, rows, malformed_rows=malformed_rows)
        return rows, {
            "format": "csv",
            "sheet": None,
            **_source_metadata(path, len(rows)),
        }

    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        selected_key: str | None = None
        array_keys: list[str] = []
        if isinstance(payload, list):
            rows = payload
        elif isinstance(payload, dict):
            arrays = [(key, value) for key, value in payload.items() if isinstance(value, list)]
            array_keys = [key for key, _ in arrays]
            if not arrays:
                raise DataContractError("invalid_json_shape", "JSON object contains no row array")
            if json_key is None and len(arrays) != 1:
                raise DataContractError(
                    "json_key_required",
                    "JSON array key is required",
                    {"array_keys": array_keys},
                )
            selected_key = json_key or arrays[0][0]
            if selected_key not in payload or not isinstance(payload[selected_key], list):
                raise DataContractError(
                    "json_key_not_found",
                    f"JSON row array not found: {selected_key}",
                )
            rows = payload[selected_key]
        else:
            raise DataContractError(
                "invalid_json_shape",
                "JSON must be a row array or contain one row array",
            )

        if not all(isinstance(row, dict) for row in rows):
            raise DataContractError("invalid_json_row", "Every JSON row must be an object")
        headers = [str(header) for header in rows[0]] if rows else []
        normalized_rows = [dict(row) for row in rows]
        _validate_table(headers, normalized_rows, first_row_number=1)
        return normalized_rows, {
            "format": "json",
            "sheet": None,
            "json_key": selected_key,
            "array_keys": array_keys,
            **_source_metadata(path, len(normalized_rows)),
        }

    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        formulas = load_workbook(path, read_only=True, data_only=False)
        try:
            available_sheets = list(workbook.sheetnames)
            if sheet is None:
                if len(available_sheets) != 1:
                    raise DataContractError(
                        "sheet_required",
                        "XLSX contains multiple worksheets",
                        {"sheets": available_sheets},
                    )
                sheet = available_sheets[0]
            if sheet not in available_sheets:
                raise DataContractError(
                    "sheet_not_found",
                    f"Worksheet not found: {sheet}",
                    {"sheets": available_sheets},
                )

            values = workbook[sheet].iter_rows(values_only=True)
            formula_values = formulas[sheet].iter_rows(values_only=True)
            try:
                header_row = next(values)
                next(formula_values)
            except StopIteration as error:
                raise DataContractError("empty_source", "source contains no rows") from error
            headers = [str(value) if value is not None else "" for value in header_row]
            rows: list[dict[str, object]] = []
            uncached_formulas: list[dict[str, object]] = []
            for row_number, (row, formula_row) in enumerate(
                zip(values, formula_values, strict=True),
                start=2,
            ):
                if len(row) != len(headers) or len(formula_row) != len(headers):
                    raise DataContractError(
                        "invalid_row_shape",
                        f"row {row_number} does not match headers",
                        {"row": row_number},
                    )
                for index, formula in enumerate(formula_row):
                    if isinstance(formula, str) and formula.startswith("=") and row[index] is None:
                        uncached_formulas.append({"row": row_number, "header": headers[index]})
                rows.append(dict(zip(headers, row, strict=True)))
            _validate_table(headers, rows)
            return rows, {
                "format": "xlsx",
                "sheet": sheet,
                "sheets": available_sheets,
                "uncached_formulas": uncached_formulas,
                **_source_metadata(path, len(rows)),
            }
        finally:
            workbook.close()
            formulas.close()

    raise DataContractError("unsupported_format", f"Unsupported source format: {suffix}")


def load_source(
    path: Path,
    sheet: str | None = None,
    json_key: str | None = None,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    """Load a source and translate parser/I/O failures into safe contract errors."""

    path = Path(path)
    details = {"source": path.name}
    if not path.exists():
        raise DataContractError(
            "source_not_found", "Source file not found", details
        )
    try:
        return _load_source(path, sheet, json_key)
    except DataContractError:
        raise
    except UnicodeDecodeError:
        raise DataContractError(
            "invalid_encoding", "Source text is not valid UTF-8", details
        ) from None
    except json.JSONDecodeError:
        raise DataContractError(
            "invalid_json", "Source JSON is malformed", details
        ) from None
    except PermissionError:
        raise DataContractError(
            "source_unreadable", "Source file is not readable", details
        ) from None
    except Exception:
        if path.suffix.casefold() == ".xlsx":
            raise DataContractError(
                "invalid_xlsx", "Source XLSX is invalid or corrupt", details
            ) from None
        raise DataContractError(
            "source_unreadable", "Source file could not be read", details
        ) from None


def parse_month_header(header: str) -> tuple[str, str] | None:
    """Return a chronological month key and stage/probability column kind."""

    normalized = _normalize_header(header)
    kind = "probability" if re.search(r"\bprobability\b", normalized) else "stage"
    without_kind = re.sub(r"\b(?:stage|status|probability)\b", " ", normalized)
    candidate = " ".join(without_kind.split())

    iso_match = _ISO_MONTH_PATTERN.fullmatch(candidate)
    if iso_match:
        year, month = (int(part) for part in iso_match.groups())
        if 1 <= month <= 12:
            return f"{year:04d}-{month:02d}", kind
        return None

    word_match = _MONTH_WORD_PATTERN.fullmatch(candidate)
    if not word_match:
        return None
    month_name, year_text = word_match.groups()
    month = MONTHS.get(month_name.casefold())
    if month is None:
        return None
    year = int(year_text)
    if len(year_text) == 2:
        year += 2000
    return f"{year:04d}-{month:02d}", kind


def _field_candidates(headers: list[str]) -> dict[str, list[str]]:
    return {
        field: [header for header in headers if _normalize_header(header) in aliases]
        for field, aliases in FIELD_ALIASES.items()
    }


def _validate_explicit_month(
    month: object,
    headers: list[str],
    seen_keys: set[str],
    used_headers: dict[str, dict[str, str]],
) -> dict[str, object]:
    if not isinstance(month, dict):
        raise DataContractError("invalid_mapping", "Every explicit month must be an object")
    required = ("key", "label", "stage")
    if any(not isinstance(month.get(field), str) or not month[field] for field in required):
        raise DataContractError(
            "invalid_mapping",
            "Explicit month requires key, label, and stage",
        )

    key = str(month["key"])
    stage = str(month["stage"])
    if key in seen_keys:
        raise DataContractError("duplicate_stage_month", f"ambiguous stage month {key}")
    if not _ISO_MONTH_PATTERN.fullmatch(key) or not 1 <= int(key[-2:]) <= 12:
        raise DataContractError("invalid_mapping", f"Invalid month key: {key}")
    if stage not in headers:
        raise DataContractError("mapping_header_not_found", f"Mapped header not found: {stage}")
    _claim_month_header(used_headers, stage, key, "stage")

    result: dict[str, object] = {"key": key, "label": month["label"], "stage": stage}
    probability = month.get("probability")
    if probability is not None:
        if not isinstance(probability, str) or probability not in headers:
            raise DataContractError(
                "mapping_header_not_found",
                f"Mapped header not found: {probability}",
            )
        _claim_month_header(used_headers, probability, key, "probability")
        result["probability"] = probability
    seen_keys.add(key)
    return result


def _claim_month_header(
    used_headers: dict[str, dict[str, str]],
    header: str,
    key: str,
    role: str,
) -> None:
    use = {"key": key, "role": role}
    if header in used_headers:
        raise DataContractError(
            "invalid_mapping",
            f"reused month header: {header}",
            {"header": header, "first": used_headers[header], "second": use},
        )
    used_headers[header] = use


def resolve_mapping(
    headers: list[str],
    explicit: dict[str, object] | None,
) -> dict[str, object]:
    """Resolve fixed aliases and chronological stage/probability columns."""

    if explicit is not None:
        if not isinstance(explicit, dict):
            raise DataContractError("invalid_mapping", "Explicit mapping must be an object")
        mapping: dict[str, object] = {}
        for field in _FIXED_FIELDS:
            header = explicit.get(field)
            if not isinstance(header, str) or not header:
                raise DataContractError("missing_mapping", f"missing field {field}")
            if header not in headers:
                raise DataContractError(
                    "mapping_header_not_found",
                    f"Mapped header not found: {header}",
                )
            mapping[field] = header
        explicit_months = explicit.get("months")
        if not isinstance(explicit_months, list):
            raise DataContractError("invalid_mapping", "Explicit mapping requires months")
        seen_keys: set[str] = set()
        used_headers: dict[str, dict[str, str]] = {}
        months = [
            _validate_explicit_month(month, headers, seen_keys, used_headers)
            for month in explicit_months
        ]
        mapping["months"] = sorted(months, key=lambda month: str(month["key"]))
        return mapping

    mapping = {}
    for field, candidates in _field_candidates(headers).items():
        if not candidates:
            raise DataContractError(
                "missing_field",
                f"missing field {field}",
                {"field": field},
            )
        if len(candidates) > 1:
            raise DataContractError(
                "ambiguous_field",
                f"ambiguous field {field}",
                {"field": field, "headers": candidates},
            )
        mapping[field] = candidates[0]

    by_month: dict[str, dict[str, list[str]]] = {}
    for header in headers:
        parsed = parse_month_header(header)
        if parsed is None:
            continue
        key, kind = parsed
        by_month.setdefault(key, {"stage": [], "probability": []})[kind].append(header)

    months: list[dict[str, object]] = []
    for key in sorted(by_month):
        stage_headers = by_month[key]["stage"]
        probability_headers = by_month[key]["probability"]
        if len(stage_headers) > 1:
            raise DataContractError(
                "duplicate_stage_month",
                f"ambiguous stage month {key}",
                {"key": key, "headers": stage_headers},
            )
        if len(probability_headers) > 1:
            raise DataContractError(
                "duplicate_probability_month",
                f"ambiguous probability month {key}",
                {"key": key, "headers": probability_headers},
            )
        if probability_headers and not stage_headers:
            raise DataContractError(
                "unpaired_probability_month",
                f"probability month has no stage: {key}",
                {"key": key, "headers": probability_headers},
            )
        if not stage_headers:
            continue
        month: dict[str, object] = {
            "key": key,
            "label": stage_headers[0],
            "stage": stage_headers[0],
        }
        if probability_headers:
            month["probability"] = probability_headers[0]
        months.append(month)
    mapping["months"] = months
    return mapping


def inspect_source(
    path: Path,
    sheet: str | None = None,
    json_key: str | None = None,
) -> dict[str, object]:
    """Return safe source metadata and deterministic mapping candidates."""

    rows, metadata = load_source(path, sheet, json_key)
    headers = list(rows[0])
    month_candidates = []
    for header in headers:
        parsed = parse_month_header(header)
        if parsed is not None:
            key, kind = parsed
            month_candidates.append({"key": key, "label": header, "kind": kind})
    month_candidates.sort(key=lambda month: (str(month["key"]), str(month["kind"])))

    ambiguities: list[dict[str, object]] = []
    mapping: dict[str, object] | None
    try:
        mapping = resolve_mapping(headers, None)
    except DataContractError as error:
        mapping = None
        ambiguities.append(
            {"code": error.code, "message": str(error), "details": error.details}
        )

    return {
        **metadata,
        "headers": headers,
        "field_candidates": _field_candidates(headers),
        "month_candidates": month_candidates,
        "mapping": mapping,
        "ambiguities": ambiguities,
    }


def _casefolded_strings(value: object, key: str) -> list[str]:
    if not isinstance(value, list) or any(
        not isinstance(item, str) or not item.strip() for item in value
    ):
        raise DataContractError("invalid_semantics", f"{key} must be a list of strings")
    folded = [item.casefold() for item in value]
    if len(set(folded)) != len(folded):
        raise DataContractError(
            "invalid_semantics", f"{key} contains a case-insensitive duplicate"
        )
    return folded


def validate_semantics(semantics: dict[str, object] | None) -> dict[str, object]:
    """Apply defaults and validate case-insensitive stage semantics."""

    if semantics is None:
        semantics = {}
    if not isinstance(semantics, dict):
        raise DataContractError("invalid_semantics", "Semantics must be an object")
    result = {
        key: list(value) if isinstance(value, list) else value
        for key, value in DEFAULT_SEMANTICS.items()
    }
    result.update(semantics)

    positive = _casefolded_strings(result["positive_terminals"], "positive_terminals")
    negative = _casefolded_strings(result["negative_terminals"], "negative_terminals")
    if set(positive) & set(negative):
        raise DataContractError(
            "invalid_semantics", "Positive and negative terminal lists overlap"
        )
    non_terminal = _casefolded_strings(
        result["non_terminal_stages"], "non_terminal_stages"
    )
    if (set(positive) | set(negative)) & set(non_terminal):
        raise DataContractError(
            "invalid_semantics", "Terminal and non-terminal stage lists overlap"
        )

    for key in ("tcv_order", "probability_order"):
        _casefolded_strings(result[key], key)

    paths = result["stage_paths"]
    if not isinstance(paths, list):
        raise DataContractError("invalid_semantics", "stage_paths must be a list")
    stage_graph: dict[str, set[str]] = {}
    for path in paths:
        if not isinstance(path, list) or not path:
            raise DataContractError(
                "invalid_semantics", "Every stage path must be a non-empty list"
            )
        folded_path = _casefolded_strings(path, "stage_paths entry")
        for previous, current in zip(folded_path, folded_path[1:]):
            stage_graph.setdefault(previous, set()).add(current)
            stage_graph.setdefault(current, set())

    visiting: set[str] = set()
    visited: set[str] = set()

    def visit(stage: str) -> bool:
        if stage in visiting:
            return True
        if stage in visited:
            return False
        visiting.add(stage)
        if any(visit(next_stage) for next_stage in stage_graph.get(stage, set())):
            return True
        visiting.remove(stage)
        visited.add(stage)
        return False

    if any(visit(stage) for stage in stage_graph if stage not in visited):
        raise DataContractError(
            "conflicting_stage_order", "Stage paths assign contradictory ranks"
        )

    transitions = result["positive_transitions"]
    if not isinstance(transitions, list):
        raise DataContractError(
            "invalid_semantics", "positive_transitions must be a list"
        )
    for transition in transitions:
        if (
            not isinstance(transition, list)
            or len(transition) != 2
            or any(not isinstance(stage, str) or not stage.strip() for stage in transition)
        ):
            raise DataContractError(
                "invalid_semantics", "Every positive transition must contain two strings"
            )
    return result


def normalize_rank(
    value: object,
    configured_order: object,
    field: str,
) -> tuple[str, int | float, str]:
    """Preserve a display value and derive a deterministic numeric sort rank."""

    if field not in {"tcv", "probability"}:
        raise ValueError(f"Unsupported rank field: {field}")
    display = str(value)
    if value is None or not display.strip() or isinstance(value, bool):
        raise DataContractError(
            f"invalid_{field}", f"Invalid {field} value: {value!r}"
        )
    if not isinstance(configured_order, list) or any(
        not isinstance(item, str) for item in configured_order
    ):
        raise DataContractError(
            "invalid_semantics", f"{field}_order must be a list of strings"
        )

    category_ranks = {
        label.casefold(): index for index, label in enumerate(configured_order)
    }
    if display.casefold() in category_ranks:
        return display, category_ranks[display.casefold()], "categorical"

    number: float | None = None
    percent = False
    multiplier = 1.0
    if isinstance(value, Real):
        number = float(value)
    elif isinstance(value, str):
        candidate = value.strip().replace(",", "").replace(" ", "")
        match = _NUMERIC_VALUE_PATTERN.fullmatch(candidate)
        if match:
            sign, numeric_text, suffix, percent_mark = match.groups()
            number = float(f"{sign}{numeric_text}")
            percent = bool(percent_mark)
            multiplier = {"k": 1_000.0, "m": 1_000_000.0, "b": 1_000_000_000.0}.get(
                (suffix or "").casefold(), 1.0
            )
    if number is None or not math.isfinite(number):
        raise DataContractError(
            f"invalid_{field}", f"Invalid {field} value: {value!r}"
        )
    rank = number * multiplier
    if field == "probability" and not percent and 0 <= rank <= 1:
        rank *= 100
    return display, rank, "numeric"


def _stage_direction(previous: str, current: str, semantics: dict[str, object]) -> str:
    previous_folded = previous.casefold()
    current_folded = current.casefold()
    if previous_folded == current_folded:
        return "neutral"

    explicit = {
        (str(start).casefold(), str(end).casefold())
        for start, end in semantics["positive_transitions"]
    }
    if (previous_folded, current_folded) in explicit:
        return "positive"

    directions: set[str] = set()
    for path in semantics["stage_paths"]:
        folded_path = [str(stage).casefold() for stage in path]
        if previous_folded in folded_path and current_folded in folded_path:
            directions.add(
                "positive"
                if folded_path.index(current_folded) > folded_path.index(previous_folded)
                else "negative"
            )
    if len(directions) > 1:
        raise DataContractError(
            "conflicting_stage_order", "Stage paths assign contradictory ranks"
        )
    return next(iter(directions), "unknown")


def classify_transition(
    previous: dict[str, object],
    current: dict[str, object],
    semantics: dict[str, object],
) -> str:
    """Classify a stage transition using terminal, stage, and probability signals."""

    validated = validate_semantics(semantics)
    current_stage = str(current.get("stage", ""))
    previous_stage = str(previous.get("stage", ""))
    current_folded = current_stage.casefold()
    if current_folded in {
        str(stage).casefold() for stage in validated["positive_terminals"]
    }:
        return "won"
    if current_folded in {
        str(stage).casefold() for stage in validated["negative_terminals"]
    }:
        return "lost"

    stage_signal = _stage_direction(previous_stage, current_stage, validated)
    probability_signal = "absent"
    previous_probability = previous.get("probability_sort")
    current_probability = current.get("probability_sort")
    if isinstance(previous_probability, Real) and isinstance(current_probability, Real):
        if current_probability > previous_probability:
            probability_signal = "positive"
        elif current_probability < previous_probability:
            probability_signal = "negative"
        else:
            probability_signal = "neutral"

    directional = {"positive", "negative"}
    if stage_signal in directional and probability_signal in directional:
        return "mixed" if stage_signal != probability_signal else stage_signal
    if stage_signal in directional:
        return stage_signal
    if probability_signal in directional:
        return probability_signal
    if stage_signal == "unknown":
        return "unknown"
    return "neutral"


def _minimal_reason(
    record: dict[str, object], code: str, message: str
) -> dict[str, object]:
    return {
        "id": record["id"],
        "source_row": record["source_row"],
        "code": code,
        "message": message,
    }


def _row_id(row: dict[str, object], source_row: int) -> str:
    for header in ("ID", "Id"):
        if header in row and str(row[header]).strip():
            return str(row[header])
    return f"ROW-{source_row:04d}"


def _terminal_kind(stage: str, semantics: dict[str, object]) -> str | None:
    folded = stage.casefold()
    if folded in {str(value).casefold() for value in semantics["positive_terminals"]}:
        return "positive"
    if folded in {str(value).casefold() for value in semantics["negative_terminals"]}:
        return "negative"
    return None


def normalize_rows(
    rows: list[dict[str, object]],
    mapping: dict[str, object],
    semantics: dict[str, object],
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    """Normalize display-safe rows and return row-local contract exclusions."""

    validated = validate_semantics(semantics)
    records: list[dict[str, object]] = []
    exclusions: list[dict[str, object]] = []
    rank_kinds: dict[str, set[str]] = {"tcv": set(), "probability": set()}

    for source_index, row in enumerate(rows):
        source_row = source_index + 2
        row_id = _row_id(row, source_row)
        reason_stub = {"id": row_id, "source_row": source_row}
        missing_field = next(
            (
                field
                for field in ("area", "sub_area", "opportunity_name")
                if row.get(str(mapping[field])) is None
                or not str(row[str(mapping[field])]).strip()
            ),
            None,
        )
        if missing_field:
            exclusions.append(
                {
                    **reason_stub,
                    "code": "missing_required_value",
                    "message": f"Missing required value: {missing_field}",
                }
            )
            continue

        try:
            tcv_display, tcv_rank, tcv_kind = normalize_rank(
                row.get(str(mapping["tcv"])), validated["tcv_order"], "tcv"
            )
            probability_display, probability_rank, probability_kind = normalize_rank(
                row.get(str(mapping["probability"])),
                validated["probability_order"],
                "probability",
            )
        except DataContractError as error:
            exclusions.append(
                {
                    **reason_stub,
                    "code": error.code,
                    "message": str(error),
                }
            )
            continue

        normalized_months: list[dict[str, object]] = []
        warnings: list[dict[str, object]] = []
        previous_populated: dict[str, object] | None = None
        blank_labels: list[str] = []
        terminal: dict[str, object] | None = None
        month_error: DataContractError | None = None

        for month in mapping["months"]:
            stage_value = row.get(str(month["stage"]))
            stage = "" if stage_value is None else str(stage_value)
            normalized_month: dict[str, object] = {
                "key": str(month["key"]),
                "label": str(month["label"]),
                "stage": stage,
                "classification": "empty",
                "skipped_months": [],
            }
            probability_header = month.get("probability")
            if probability_header is not None:
                month_probability = row.get(str(probability_header))
                if month_probability is not None and str(month_probability).strip():
                    try:
                        display, rank, _ = normalize_rank(
                            month_probability,
                            validated["probability_order"],
                            "probability",
                        )
                    except DataContractError as error:
                        month_error = error
                        break
                    normalized_month["probability_display"] = display
                    normalized_month["probability_sort"] = rank

            if not stage.strip():
                blank_labels.append(str(month["label"]))
                normalized_months.append(normalized_month)
                continue

            terminal_kind = _terminal_kind(stage, validated)
            if previous_populated is None:
                classification = (
                    "won"
                    if terminal_kind == "positive"
                    else "lost" if terminal_kind == "negative" else "initial"
                )
            else:
                classification = classify_transition(
                    previous_populated, normalized_month, validated
                )
            normalized_month["classification"] = classification
            if blank_labels and previous_populated is not None:
                normalized_month["skipped_months"] = list(blank_labels)
                warnings.append(
                    {
                        "code": "skipped_blank_months",
                        "message": "Transition skips blank months",
                        "month": normalized_month["key"],
                        "skipped_months": list(blank_labels),
                    }
                )
            blank_labels.clear()
            if classification == "mixed":
                warnings.append(
                    {
                        "code": "mixed_signals",
                        "message": "Stage and probability signals conflict",
                        "month": normalized_month["key"],
                    }
                )
            elif classification == "unknown":
                warnings.append(
                    {
                        "code": "unknown_transition",
                        "message": "Transition stage order is unknown",
                        "month": normalized_month["key"],
                    }
                )
            normalized_months.append(normalized_month)
            previous_populated = normalized_month
            if terminal_kind is not None:
                terminal = {
                    "kind": terminal_kind,
                    "month": normalized_month["key"],
                    "index": len(normalized_months) - 1,
                }
                break

        if month_error is not None:
            exclusions.append(
                {
                    **reason_stub,
                    "code": month_error.code,
                    "message": str(month_error),
                }
            )
            continue

        populated_count = sum(bool(str(month["stage"]).strip()) for month in normalized_months)
        if populated_count < 2:
            exclusions.append(
                {
                    **reason_stub,
                    "code": "insufficient_stages",
                    "message": "At least two populated stages are required",
                }
            )
            continue

        rank_kinds["tcv"].add(tcv_kind)
        rank_kinds["probability"].add(probability_kind)
        records.append(
            {
                "id": row_id,
                "source_row": source_row,
                "area": str(row[str(mapping["area"])]),
                "sub_area": str(row[str(mapping["sub_area"])]),
                "opportunity_name": str(row[str(mapping["opportunity_name"])]),
                "tcv": {"display": tcv_display, "sort": tcv_rank, "kind": tcv_kind},
                "probability": {
                    "display": probability_display,
                    "sort": probability_rank,
                    "kind": probability_kind,
                },
                "months": normalized_months,
                "terminal": terminal,
                "warnings": warnings,
            }
        )

    for field, kinds in rank_kinds.items():
        if len(kinds) > 1:
            raise DataContractError(
                f"mixed_{field}_types",
                f"{field.upper() if field == 'tcv' else field.capitalize()} values mix numeric and categorical ranks",
            )
    return records, exclusions


def _validate_filters(filters: dict[str, object] | None) -> dict[str, object]:
    if filters is None:
        return {}
    if not isinstance(filters, dict):
        raise DataContractError("invalid_filters", "Filters must be an object")
    unknown = sorted(set(filters) - _FILTER_KEYS)
    if unknown:
        raise DataContractError(
            "invalid_filters", "Unknown filter keys", {"keys": unknown}
        )
    for key in ("areas", "sub_areas"):
        value = filters.get(key)
        if value not in (None, []) and (
            not isinstance(value, list) or any(not isinstance(item, str) for item in value)
        ):
            raise DataContractError("invalid_filters", f"{key} must be a list of strings")
    contains = filters.get("opportunity_contains")
    if contains is not None and not isinstance(contains, str):
        raise DataContractError(
            "invalid_filters", "opportunity_contains must be a string"
        )
    for prefix in ("tcv", "probability"):
        minimum = filters.get(f"{prefix}_min")
        maximum = filters.get(f"{prefix}_max")
        if minimum is not None and not isinstance(minimum, Real):
            raise DataContractError("invalid_filters", f"{prefix}_min must be numeric")
        if maximum is not None and not isinstance(maximum, Real):
            raise DataContractError("invalid_filters", f"{prefix}_max must be numeric")
        if minimum is not None and maximum is not None and minimum > maximum:
            raise DataContractError(
                "invalid_filters", f"{prefix}_min must not exceed {prefix}_max"
            )
    return dict(filters)


def apply_filters(
    records: list[dict[str, object]], filters: dict[str, object] | None
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    """Apply deterministic fixed-field filters without leaking excluded row data."""

    validated = _validate_filters(filters)
    areas = {value.casefold() for value in validated.get("areas", [])}
    sub_areas = {value.casefold() for value in validated.get("sub_areas", [])}
    contains = str(validated.get("opportunity_contains", "")).casefold()
    included: list[dict[str, object]] = []
    excluded: list[dict[str, object]] = []

    for record in records:
        matches = (
            (not areas or str(record["area"]).casefold() in areas)
            and (not sub_areas or str(record["sub_area"]).casefold() in sub_areas)
            and (not contains or contains in str(record["opportunity_name"]).casefold())
        )
        for field in ("tcv", "probability"):
            rank = record[field]["sort"]
            minimum = validated.get(f"{field}_min")
            maximum = validated.get(f"{field}_max")
            matches = matches and (minimum is None or rank >= minimum)
            matches = matches and (maximum is None or rank <= maximum)
        if matches:
            included.append(record)
        else:
            excluded.append(
                _minimal_reason(record, "filter_not_matched", "Row does not match filters")
            )
    return included, excluded


def populated_stage_count(record: dict[str, object]) -> int:
    return sum(bool(str(month["stage"]).strip()) for month in record["months"])


def _record_sort_key(record: dict[str, object]) -> tuple[object, ...]:
    return (
        str(record["area"]).casefold(),
        str(record["sub_area"]).casefold(),
        -float(record["tcv"]["sort"]),
        -float(record["probability"]["sort"]),
        str(record["opportunity_name"]).casefold(),
        int(record["source_row"]),
    )


def select_records(
    records: list[dict[str, object]], view: str
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    """Select a canonical view and sort included records deterministically."""

    predicates = {
        "wins": lambda record: record["terminal"] is not None
        and record["terminal"]["kind"] == "positive",
        "losses": lambda record: record["terminal"] is not None
        and record["terminal"]["kind"] == "negative",
        "all-progression": lambda record: populated_stage_count(record) >= 2,
        "positive-progression": lambda record: any(
            month["classification"] in {"positive", "won"}
            for month in record["months"]
        ),
    }
    if view not in predicates:
        raise DataContractError("unsupported_view", f"Unsupported view: {view}")
    included = [record for record in records if predicates[view](record)]
    excluded = [
        _minimal_reason(record, "view_not_matched", f"Row does not match view: {view}")
        for record in records
        if not predicates[view](record)
    ]
    included.sort(key=_record_sort_key)
    return included, excluded
